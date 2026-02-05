#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phoenix Trading Excel Template Validator
검증 항목: B12-B22 필드 완전성 및 정합성
"""

import openpyxl
import sys
import re
import os
from datetime import datetime
from pathlib import Path

# Windows 콘솔 UTF-8 인코딩 강제
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

class ExcelValidator:
    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath)
        self.ws = self.wb.active
        self.results = []
        self.errors = []
        self.warnings = []

    def validate_all(self):
        """전체 검증 실행"""
        print("=" * 60)
        print("Phoenix Trading Excel 템플릿 검증 시작")
        print("=" * 60)
        print(f"파일: {self.filepath}")
        print(f"시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        # 1. 필수 인증 정보
        self._validate_kis_credentials()

        # 2. Tier 매도가
        self._validate_tier_prices()

        # 3. 추가 설정
        self._validate_additional_settings()

        # 결과 출력
        self._print_results()

        # 리포트 생성
        self._generate_report()

        return len(self.errors) == 0

    def _validate_kis_credentials(self):
        """KIS API 인증 정보 검증"""
        print("\n[1] KIS API 인증 정보 검증")
        print("-" * 60)

        # B12: APP KEY
        app_key = self.ws['B12'].value
        if not app_key:
            self.errors.append("B12: KIS APP KEY가 비어있습니다")
            print("❌ B12: KIS APP KEY - 없음")
        elif len(str(app_key)) < 30:
            self.warnings.append("B12: APP KEY 길이가 너무 짧습니다 (의심)")
            print(f"⚠️  B12: KIS APP KEY - {app_key[:10]}... (길이: {len(app_key)})")
        else:
            print(f"✅ B12: KIS APP KEY - {app_key[:10]}... (길이: {len(app_key)})")

        # B13: APP SECRET
        app_secret = self.ws['B13'].value
        if not app_secret:
            self.errors.append("B13: KIS APP SECRET이 비어있습니다")
            print("❌ B13: KIS APP SECRET - 없음")
        elif app_secret == "your_kis_app_secret_here":
            self.errors.append("B13: 기본값(placeholder)이 남아있습니다")
            print("❌ B13: KIS APP SECRET - 기본값 (수정 필요)")
        else:
            print(f"✅ B13: KIS APP SECRET - {app_secret[:10]}...")

        # B14: 계좌번호
        account = self.ws['B14'].value
        if not account:
            self.errors.append("B14: 계좌번호가 비어있습니다")
            print("❌ B14: 계좌번호 - 없음")
        else:
            # 형식: 12345678-01
            pattern = r'^\d{8}-\d{2}$'
            if re.match(pattern, str(account)):
                print(f"✅ B14: 계좌번호 - {account}")
            else:
                self.warnings.append(f"B14: 계좌번호 형식이 이상합니다 ({account})")
                print(f"⚠️  B14: 계좌번호 - {account} (형식 확인 필요)")

        # B15: 시스템 가동
        is_active = self.ws['B15'].value
        if is_active is True or str(is_active).upper() == "TRUE":
            print(f"✅ B15: 시스템 가동 - {is_active} (활성화)")
        elif is_active is False or str(is_active).upper() == "FALSE":
            self.warnings.append("B15: 시스템이 비활성화 상태입니다")
            print(f"⚠️  B15: 시스템 가동 - {is_active} (비활성화)")
        else:
            self.errors.append(f"B15: 시스템 가동 값이 유효하지 않습니다 ({is_active})")
            print(f"❌ B15: 시스템 가동 - {is_active} (유효하지 않음)")

    def _validate_tier_prices(self):
        """Tier 매도가 검증"""
        print("\n[2] Tier 매도가 검증")
        print("-" * 60)

        tier_cells = {
            'B16': 'Tier 1 매도가',
            'B17': 'Tier 2 매도가',
            'B18': 'Tier 3 매도가',
            'B19': 'Tier 4 매도가',
            'B20': 'Tier 5 매도가'
        }

        tier_prices = []
        for cell, label in tier_cells.items():
            value = self.ws[cell].value

            if cell == 'B16':  # Tier 1은 필수
                if value is None or value == 0:
                    self.errors.append(f"{cell}: {label}이 설정되지 않았습니다 (필수)")
                    print(f"❌ {cell}: {label} - 없음 (필수)")
                elif isinstance(value, (int, float)) and value > 0:
                    tier_prices.append(value)
                    print(f"✅ {cell}: {label} - ${value}")
                else:
                    self.errors.append(f"{cell}: {label} 값이 유효하지 않습니다 ({value})")
                    print(f"❌ {cell}: {label} - {value} (유효하지 않음)")
            else:  # Tier 2-5는 선택
                if value is None or value == 0 or str(value).upper() == "FALSE":
                    print(f"⚪ {cell}: {label} - 미사용")
                elif isinstance(value, (int, float)) and value > 0:
                    tier_prices.append(value)
                    print(f"✅ {cell}: {label} - ${value}")
                else:
                    self.warnings.append(f"{cell}: {label} 값이 이상합니다 ({value})")
                    print(f"⚠️  {cell}: {label} - {value}")

        # Tier 가격 순서 검증 (오름차순이어야 함)
        if len(tier_prices) > 1:
            for i in range(len(tier_prices) - 1):
                if tier_prices[i] >= tier_prices[i+1]:
                    self.warnings.append(f"Tier 매도가 순서가 올바르지 않습니다: ${tier_prices[i]} >= ${tier_prices[i+1]}")
                    print(f"⚠️  경고: Tier 가격이 오름차순이 아닙니다!")

    def _validate_additional_settings(self):
        """추가 설정 검증"""
        print("\n[3] 추가 설정 검증")
        print("-" * 60)

        # B22: Tier 1 매수%
        tier1_buy_pct = self.ws['B22'].value
        if tier1_buy_pct is None or tier1_buy_pct == 0:
            self.errors.append("B22: Tier 1 매수%가 설정되지 않았습니다")
            print("❌ B22: Tier 1 매수% - 없음")
        elif isinstance(tier1_buy_pct, (int, float)) and 0 < tier1_buy_pct <= 100:
            print(f"✅ B22: Tier 1 매수% - {tier1_buy_pct}%")
        else:
            self.warnings.append(f"B22: Tier 1 매수% 값이 이상합니다 ({tier1_buy_pct})")
            print(f"⚠️  B22: Tier 1 매수% - {tier1_buy_pct}")

    def _print_results(self):
        """검증 결과 출력"""
        print("\n" + "=" * 60)
        print("검증 결과 요약")
        print("=" * 60)

        total_checks = len(self.errors) + len(self.warnings)

        if len(self.errors) == 0 and len(self.warnings) == 0:
            print("✅ 모든 검증 통과! 실거래 준비 완료")
            print("\n🚀 다음 단계:")
            print("   1. B15를 TRUE로 변경 (시스템 가동)")
            print("   2. phoenix_main.py 실행")
            print("   3. 로그 모니터링")
        else:
            if len(self.errors) > 0:
                print(f"\n❌ 에러: {len(self.errors)}개")
                for i, error in enumerate(self.errors, 1):
                    print(f"   {i}. {error}")

            if len(self.warnings) > 0:
                print(f"\n⚠️  경고: {len(self.warnings)}개")
                for i, warning in enumerate(self.warnings, 1):
                    print(f"   {i}. {warning}")

            print("\n⛔ 실거래 전 반드시 수정이 필요합니다!")

        print("=" * 60)

    def _generate_report(self):
        """검증 리포트 생성"""
        report_dir = Path('.claude/logs')
        report_dir.mkdir(parents=True, exist_ok=True)

        report_path = report_dir / 'Excel-Validation-Report.md'

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(f"# Phoenix Trading Excel 검증 리포트\n\n")
            f.write(f"**검증 시간:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"**파일:** {self.filepath}\n\n")

            f.write("## 검증 결과\n\n")

            if len(self.errors) == 0 and len(self.warnings) == 0:
                f.write("### ✅ 상태: 통과\n\n")
                f.write("모든 검증 항목을 통과했습니다. 실거래 준비 완료.\n\n")
            else:
                f.write("### ⚠️ 상태: 수정 필요\n\n")

                if len(self.errors) > 0:
                    f.write(f"#### ❌ 에러 ({len(self.errors)}개)\n\n")
                    for i, error in enumerate(self.errors, 1):
                        f.write(f"{i}. {error}\n")
                    f.write("\n")

                if len(self.warnings) > 0:
                    f.write(f"#### ⚠️ 경고 ({len(self.warnings)}개)\n\n")
                    for i, warning in enumerate(self.warnings, 1):
                        f.write(f"{i}. {warning}\n")
                    f.write("\n")

            f.write("## 체크리스트\n\n")
            f.write("- [ ] B12: KIS APP KEY 입력\n")
            f.write("- [ ] B13: KIS APP SECRET 입력\n")
            f.write("- [ ] B14: 계좌번호 입력\n")
            f.write("- [ ] B15: 시스템 가동 TRUE로 변경\n")
            f.write("- [ ] B16: Tier 1 매도가 설정\n")
            f.write("- [ ] B22: Tier 1 매수% 설정\n")
            f.write("\n")
            f.write("---\n")
            f.write("*자동 생성된 리포트*\n")

        print(f"\n📄 리포트 저장됨: {report_path}")


def main():
    if len(sys.argv) < 2:
        print("사용법: python excel_validator.py <excel_file>")
        sys.exit(1)

    filepath = sys.argv[1]

    if not Path(filepath).exists():
        print(f"❌ 파일을 찾을 수 없습니다: {filepath}")
        sys.exit(1)

    validator = ExcelValidator(filepath)
    success = validator.validate_all()

    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
