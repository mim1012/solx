#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KIS API Health Check Script
검증 항목: 토큰 발급, 계좌 조회, API 응답 시간
"""

import requests
import openpyxl
import sys
import time
import json
from datetime import datetime
from pathlib import Path

# Windows 콘솔 UTF-8 인코딩 강제
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


class KISHealthChecker:
    BASE_URL = "https://openapi.koreainvestment.com:9443"

    def __init__(self):
        self.app_key = None
        self.app_secret = None
        self.account_no = None
        self.access_token = None
        self.results = []
        self.errors = []
        self.warnings = []

    def run(self):
        """전체 헬스체크 실행"""
        print("=" * 60)
        print("KIS API 헬스체크 시작")
        print("=" * 60)
        print(f"시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        # 1. Excel 설정 읽기
        if not self._load_config_from_excel():
            self._print_results()
            return False

        # 2. API 연결 테스트
        self._test_api_connection()

        # 3. 토큰 발급 테스트
        if self._test_token_issuance():
            # 4. 계좌 조회 테스트 (토큰 발급 성공 시에만)
            self._test_account_query()

        # 결과 출력
        self._print_results()

        # 리포트 생성
        self._generate_report()

        return len(self.errors) == 0

    def _load_config_from_excel(self):
        """Excel에서 설정 읽기"""
        print("\n[1] Excel 설정 로드")
        print("-" * 60)

        try:
            excel_path = "phoenix_grid_template_v3.xlsx"
            if not Path(excel_path).exists():
                self.errors.append(f"Excel 파일을 찾을 수 없습니다: {excel_path}")
                print(f"❌ Excel 파일 없음: {excel_path}")
                return False

            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            # B12: APP KEY
            self.app_key = ws['B12'].value
            if not self.app_key:
                self.errors.append("B12: KIS APP KEY가 비어있습니다")
                print("❌ B12: APP KEY - 없음")
                return False
            else:
                masked_key = f"{self.app_key[:8]}...{self.app_key[-4:]}" if len(self.app_key) > 12 else "***"
                print(f"✅ B12: APP KEY - {masked_key}")

            # B13: APP SECRET
            self.app_secret = ws['B13'].value
            if not self.app_secret:
                self.errors.append("B13: KIS APP SECRET이 비어있습니다")
                print("❌ B13: APP SECRET - 없음")
                return False
            elif self.app_secret == "your_kis_app_secret_here":
                self.errors.append("B13: 기본값(placeholder)이 남아있습니다")
                print("❌ B13: APP SECRET - 기본값")
                return False
            else:
                print(f"✅ B13: APP SECRET - {self.app_secret[:8]}...")

            # B14: 계좌번호
            self.account_no = ws['B14'].value
            if not self.account_no:
                self.warnings.append("B14: 계좌번호가 비어있습니다 (계좌 조회 스킵)")
                print("⚠️  B14: 계좌번호 - 없음 (계좌 조회 스킵)")
            else:
                # 계좌번호 마스킹 (예: 12345678-01 → 1234****-01)
                masked_account = f"{self.account_no[:4]}****{self.account_no[-3:]}"
                print(f"✅ B14: 계좌번호 - {masked_account}")

            return True

        except Exception as e:
            self.errors.append(f"Excel 읽기 실패: {e}")
            print(f"❌ Excel 읽기 실패: {e}")
            return False

    def _test_api_connection(self):
        """API 기본 연결 테스트"""
        print("\n[2] API 연결 테스트")
        print("-" * 60)

        try:
            start_time = time.time()
            response = requests.get(f"{self.BASE_URL}", timeout=5)
            elapsed = time.time() - start_time

            if response.status_code == 200 or response.status_code == 404:
                # 404도 연결은 성공 (엔드포인트가 존재하지 않을 뿐)
                print(f"✅ API 서버 연결 성공 ({elapsed:.2f}초)")
            else:
                self.warnings.append(f"API 서버 응답 이상: {response.status_code}")
                print(f"⚠️  API 서버 응답: {response.status_code} ({elapsed:.2f}초)")

        except requests.exceptions.Timeout:
            self.errors.append("API 서버 응답 시간 초과 (5초)")
            print("❌ API 서버 타임아웃 (네트워크 확인 필요)")
        except Exception as e:
            self.errors.append(f"API 연결 실패: {e}")
            print(f"❌ API 연결 실패: {e}")

    def _test_token_issuance(self):
        """토큰 발급 테스트"""
        print("\n[3] 토큰 발급 테스트")
        print("-" * 60)

        url = f"{self.BASE_URL}/oauth2/tokenP"
        headers = {"content-type": "application/json"}
        body = {
            "grant_type": "client_credentials",
            "appkey": self.app_key,
            "appsecret": self.app_secret
        }

        try:
            start_time = time.time()
            response = requests.post(url, headers=headers, json=body, timeout=10)
            elapsed = time.time() - start_time

            if response.status_code == 200:
                data = response.json()
                self.access_token = data.get("access_token")

                if self.access_token:
                    expires_in = data.get("expires_in", 0)
                    print(f"✅ 토큰 발급 성공 ({elapsed:.2f}초)")
                    print(f"   - 토큰: {self.access_token[:20]}...")
                    print(f"   - 만료: {expires_in}초 후")
                    return True
                else:
                    self.errors.append("토큰 응답에 access_token이 없습니다")
                    print(f"❌ 토큰 발급 실패: access_token 없음")
                    return False
            else:
                error_msg = response.json().get("msg1", response.text)
                self.errors.append(f"토큰 발급 실패 ({response.status_code}): {error_msg}")
                print(f"❌ 토큰 발급 실패 ({response.status_code})")
                print(f"   - 응답: {error_msg}")
                return False

        except requests.exceptions.Timeout:
            self.errors.append("토큰 발급 타임아웃 (10초)")
            print("❌ 토큰 발급 타임아웃")
            return False
        except Exception as e:
            self.errors.append(f"토큰 발급 오류: {e}")
            print(f"❌ 토큰 발급 오류: {e}")
            return False

    def _test_account_query(self):
        """계좌 조회 테스트"""
        if not self.account_no:
            print("\n[4] 계좌 조회 테스트")
            print("-" * 60)
            print("⚪ 계좌번호 없음 - 스킵")
            return

        print("\n[4] 계좌 조회 테스트")
        print("-" * 60)

        # 계좌번호 파싱 (12345678-01 → 12345678, 01)
        try:
            parts = self.account_no.split('-')
            if len(parts) != 2:
                self.warnings.append("계좌번호 형식이 올바르지 않습니다")
                print(f"⚠️  계좌번호 형식 오류: {self.account_no}")
                return

            account_base = parts[0]
            account_suffix = parts[1]

            # 해외주식 잔고 조회 API (CTRP6548R)
            url = f"{self.BASE_URL}/uapi/overseas-stock/v1/trading/inquire-balance"
            headers = {
                "content-type": "application/json",
                "authorization": f"Bearer {self.access_token}",
                "appkey": self.app_key,
                "appsecret": self.app_secret,
                "tr_id": "CTRP6548R",
                "custtype": "P"  # 개인
            }
            params = {
                "CANO": account_base,
                "ACNT_PRDT_CD": account_suffix,
                "OVRS_EXCG_CD": "NASD",  # NASDAQ
                "TR_CRCY_CD": "USD",
                "CTX_AREA_FK200": "",
                "CTX_AREA_NK200": ""
            }

            start_time = time.time()
            response = requests.get(url, headers=headers, params=params, timeout=10)
            elapsed = time.time() - start_time

            if response.status_code == 200:
                data = response.json()
                rt_cd = data.get("rt_cd")

                if rt_cd == "0":
                    print(f"✅ 계좌 조회 성공 ({elapsed:.2f}초)")

                    # 잔고 정보 출력 (선택)
                    output1 = data.get("output1", [])
                    if len(output1) > 0:
                        print(f"   - 보유 종목 수: {len(output1)}개")
                    else:
                        print("   - 보유 종목: 없음")
                else:
                    error_msg = data.get("msg1", "알 수 없는 오류")
                    self.warnings.append(f"계좌 조회 실패: {error_msg}")
                    print(f"⚠️  계좌 조회 실패: {error_msg}")
            else:
                error_msg = response.json().get("msg1", response.text)
                self.errors.append(f"계좌 조회 HTTP 오류 ({response.status_code}): {error_msg}")
                print(f"❌ 계좌 조회 실패 ({response.status_code})")
                print(f"   - 응답: {error_msg}")

        except Exception as e:
            self.errors.append(f"계좌 조회 오류: {e}")
            print(f"❌ 계좌 조회 오류: {e}")

    def _print_results(self):
        """검증 결과 출력"""
        print("\n" + "=" * 60)
        print("헬스체크 결과 요약")
        print("=" * 60)

        if len(self.errors) == 0 and len(self.warnings) == 0:
            print("✅ 모든 검증 통과! KIS API 연결 정상")
            print("\n🚀 다음 단계:")
            print("   1. Excel B15를 TRUE로 변경 (시스템 가동)")
            print("   2. phoenix_main.py 실행")
            print("   3. 실시간 시세 및 주문 모니터링")
        else:
            if len(self.errors) > 0:
                print(f"\n❌ 에러: {len(self.errors)}개")
                for i, error in enumerate(self.errors, 1):
                    print(f"   {i}. {error}")

            if len(self.warnings) > 0:
                print(f"\n⚠️  경고: {len(self.warnings)}개")
                for i, warning in enumerate(self.warnings, 1):
                    print(f"   {i}. {warning}")

            print("\n⛔ API 연결 문제 해결 후 다시 시도하세요!")

        print("=" * 60)

    def _generate_report(self):
        """헬스체크 리포트 생성"""
        report_dir = Path('.claude/logs')
        report_dir.mkdir(parents=True, exist_ok=True)

        report_path = report_dir / 'KIS-Health-Report.md'

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(f"# KIS API 헬스체크 리포트\n\n")
            f.write(f"**검증 시간:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            f.write("## 검증 결과\n\n")

            if len(self.errors) == 0 and len(self.warnings) == 0:
                f.write("### ✅ 상태: 정상\n\n")
                f.write("모든 검증 항목을 통과했습니다. KIS API 연결이 정상입니다.\n\n")
            else:
                f.write("### ⚠️ 상태: 문제 발견\n\n")

                if len(self.errors) > 0:
                    f.write(f"#### ❌ 에러 ({len(self.errors)}개)\n\n")
                    for i, error in enumerate(self.errors, 1):
                        f.write(f"{i}. {error}\n")
                    f.write("\n")

                if len(self.warnings) > 0:
                    f.write(f"#### ⚠️ 경고 ({len(self.warnings)}개)\n\n")
                    for i, warning in enumerate(self.warnings, 1):
                        f.write(f"{i}. {warning}\n")
                    f.write("\n")

            f.write("## 체크리스트\n\n")
            f.write("- [ ] B12: KIS APP KEY 유효\n")
            f.write("- [ ] B13: KIS APP SECRET 유효\n")
            f.write("- [ ] 토큰 발급 성공\n")
            f.write("- [ ] 계좌 조회 성공\n")
            f.write("- [ ] API 응답 시간 < 2초\n")
            f.write("\n")
            f.write("---\n")
            f.write("*자동 생성된 리포트*\n")

        print(f"\n📄 리포트 저장됨: {report_path}")


def main():
    checker = KISHealthChecker()
    success = checker.run()
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
