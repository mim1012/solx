"""
KIS API 문서 분석 스크립트
- 다운로드한 Excel 파일에서 API 스펙 추출
- 현재 구현과 비교
"""

import openpyxl
import json
from pathlib import Path

# KIS API 문서 파일
KIS_DOC_PATH = r"C:\Users\PC_1M\Downloads\[해외주식] 주문_계좌.xlsx"


def analyze_kis_api_doc():
    """KIS API 문서 분석"""

    print("=" * 80)
    print("KIS API 문서 분석")
    print("=" * 80)

    if not Path(KIS_DOC_PATH).exists():
        print(f"[!] File not found: {KIS_DOC_PATH}")
        return

    try:
        wb = openpyxl.load_workbook(KIS_DOC_PATH, data_only=True)

        print(f"\n[FILE] {KIS_DOC_PATH}")
        print(f"[SHEETS] {wb.sheetnames}")
        print()

        # 각 시트 분석
        for sheet_name in wb.sheetnames:
            print("=" * 80)
            print(f"시트: {sheet_name}")
            print("=" * 80)

            ws = wb[sheet_name]

            # 시트의 첫 20행 출력 (구조 파악)
            print("\n[처음 20행 미리보기]")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
                # 빈 행 건너뛰기
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                # 행 출력 (첫 10개 컬럼만)
                row_data = [str(cell)[:50] if cell is not None else "" for cell in row[:10]]
                print(f"Row {row_idx:2d}: {' | '.join(row_data)}")

            print()

            # 특정 키워드 검색 (API 파라미터 찾기)
            print(f"\n['{sheet_name}' 시트에서 주요 키워드 검색]")
            keywords_found = {}

            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        cell_lower = cell.lower()

                        # 잔고 관련
                        if "잔고" in cell or "balance" in cell_lower:
                            keywords_found.setdefault("잔고조회", []).append(cell)

                        # 주문 관련
                        if "주문" in cell or "order" in cell_lower:
                            keywords_found.setdefault("주문", []).append(cell)

                        # 파라미터 필드명
                        if any(key in cell.upper() for key in ["WCRC_FRCR_DVSN_CD", "CANO", "ACNT_PRDT_CD", "NATN_CD", "TR_MKET_CD", "INQR_DVSN_CD"]):
                            keywords_found.setdefault("파라미터", []).append(cell)

            for keyword, items in keywords_found.items():
                print(f"\n  [{keyword}] 발견 ({len(items)}건):")
                for item in set(items[:5]):  # 중복 제거, 최대 5개
                    print(f"    - {item}")

            print()

        wb.close()

    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()


def check_current_implementation():
    """현재 구현된 파라미터 확인"""

    print("\n" + "=" * 80)
    print("현재 구현 파라미터 확인")
    print("=" * 80)

    from src.kis_rest_adapter import KisRestAdapter

    print("\n[잔고조회 API - get_balance()]")
    print("URL: /uapi/overseas-stock/v1/trading/inquire-present-balance")
    print("TR_ID: CTRP6504R")
    print("파라미터:")
    print("  - CANO: 계좌번호 (8자리)")
    print("  - ACNT_PRDT_CD: 계좌상품코드 (2자리)")
    print("  - WCRC_FRCR_DVSN_CD: '02' (외화)")
    print("  - NATN_CD: '840' (미국)")
    print("  - TR_MKET_CD: '00' (전체)")
    print("  - INQR_DVSN_CD: '00' (전체)")

    print("\n[주문 API - send_order()]")
    print("URL: /uapi/overseas-stock/v1/trading/order")
    print("TR_ID: JTTT1002U")
    print("파라미터:")
    print("  - CANO: 계좌번호 (8자리)")
    print("  - ACNT_PRDT_CD: 계좌상품코드 (2자리)")
    print("  - OVRS_EXCG_CD: 'NASD' (나스닥)")
    print("  - PDNO: 종목코드")
    print("  - ORD_QTY: 주문수량")
    print("  - OVRS_ORD_UNPR: 주문단가")
    print("  - ORD_SVR_DVSN_CD: '0'")
    print("  - ORD_DVSN: '00' (지정가) / '01' (시장가)")


if __name__ == "__main__":
    analyze_kis_api_doc()
    check_current_implementation()
