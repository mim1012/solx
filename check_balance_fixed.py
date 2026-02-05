"""
한국투자증권 USD 예수금 조회 (올바른 API 사용)
"""
import sys
import io
import requests
import json
from datetime import datetime
import openpyxl

# UTF-8 출력 설정
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

def get_access_token(app_key, app_secret):
    """OAuth 토큰 발급"""
    url = "https://openapi.koreainvestment.com:9443/oauth2/tokenP"
    headers = {"content-type": "application/json"}
    body = {
        "grant_type": "client_credentials",
        "appkey": app_key,
        "appsecret": app_secret
    }

    try:
        response = requests.post(url, headers=headers, data=json.dumps(body), timeout=10)
        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            print(f"[ERROR] Token failed: {response.status_code}")
            return None
    except Exception as e:
        print(f"[ERROR] Token error: {e}")
        return None

def get_usd_balance(app_key, app_secret, account_no, access_token):
    """USD 예수금 조회 (올바른 API)"""
    # 계좌번호 파싱
    if '-' in account_no:
        cano, acnt_prdt_cd = account_no.split('-')
    else:
        cano = account_no[:8]
        acnt_prdt_cd = account_no[8:]

    # 올바른 API: 해외주식 잔고조회
    url = "https://openapi.koreainvestment.com:9443/uapi/overseas-stock/v1/trading/inquire-balance"

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "TTTS3012R",  # 해외주식 잔고조회
        "custtype": "P"
    }

    params = {
        "CANO": cano,
        "ACNT_PRDT_CD": acnt_prdt_cd,
        "OVRS_EXCG_CD": "NASD",
        "TR_CRCY_CD": "USD",
        "CTX_AREA_FK200": "",
        "CTX_AREA_NK200": ""
    }

    try:
        response = requests.get(url, headers=headers, params=params, timeout=10)

        if response.status_code == 200:
            data = response.json()
            rt_cd = data.get("rt_cd")

            if rt_cd == "0":
                output1 = data.get("output1", [])
                output2 = data.get("output2")

                # USD 예수금 추출
                cash = 0.0
                if output2:
                    if isinstance(output2, dict):
                        cash = float(output2.get("frcr_drwg_psbl_amt_1", 0) or 0)
                    elif isinstance(output2, list) and len(output2) > 0:
                        if isinstance(output2[0], dict):
                            cash = float(output2[0].get("frcr_drwg_psbl_amt_1", 0) or 0)

                return cash, data, output1
            else:
                print(f"[ERROR] Query failed: rt_cd={rt_cd}, msg={data.get('msg1')}")
                return None, data, []
        else:
            print(f"[ERROR] HTTP error: {response.status_code}")
            return None, None, []
    except Exception as e:
        print(f"[ERROR] Balance query error: {e}")
        return None, None, []

def main():
    print("=" * 60)
    print("Korea Investment USD Balance Check (Fixed)")
    print("=" * 60)
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Excel에서 설정 읽기
    excel_path = 'phoenix_grid_template_v3.xlsx'

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        app_key = ws['B12'].value
        app_secret = ws['B13'].value
        account_no = ws['B14'].value

        print(f"[INFO] Reading from Excel: {excel_path}")
        print(f"Account: {account_no}")
        print(f"APP KEY: {app_key[:20]}..." if app_key else "APP KEY: None")
        print()

        if not app_key or not app_secret or not account_no:
            print("[ERROR] Excel settings incomplete!")
            print("Please check B12, B13, B14")
            return

    except FileNotFoundError:
        print(f"[ERROR] Excel file not found: {excel_path}")
        print("Run this in the 'release' folder.")
        return
    except Exception as e:
        print(f"[ERROR] Excel read failed: {e}")
        return

    # 토큰 발급
    print("Requesting OAuth token...")
    access_token = get_access_token(app_key, app_secret)

    if not access_token:
        print("[ERROR] Token issuance failed")
        return

    print("[OK] Token issued successfully")
    print()

    # USD 예수금 조회
    print("Querying USD balance (using correct API)...")
    usd_balance, data, holdings = get_usd_balance(app_key, app_secret, account_no, access_token)

    print()
    print("=" * 60)

    if usd_balance is not None:
        print(f"[RESULT] USD Balance: ${usd_balance:,.2f}")

        if usd_balance == 0:
            print()
            print("[WARNING] USD balance is $0.00")
        else:
            print()
            print("[SUCCESS] Tradeable USD balance confirmed!")

        # 보유 종목 표시
        if holdings and len(holdings) > 0:
            print()
            print("Holdings:")
            for item in holdings:
                ticker = item.get("ovrs_pdno", "")
                qty = item.get("ovrs_cblc_qty", "0")
                if ticker:
                    print(f"  {ticker}: {qty} shares")
        else:
            print()
            print("Holdings: None")

        # API 응답 상세
        if data:
            print()
            print("-" * 60)
            print("API Response:")
            print(f"  rt_cd: {data.get('rt_cd')}")
            print(f"  msg1: {data.get('msg1', '').strip()}")
            print()
            print("Full API Response (DEBUG):")
            print(json.dumps(data, indent=2, ensure_ascii=False))
    else:
        print("[ERROR] Balance query failed")

    print("=" * 60)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nTerminated.")
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()

    print()
    input("Press Enter to exit...")
