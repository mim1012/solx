"""
한국투자증권 USD 예수금 조회 (실제 Phoenix Trading 방식)
"""
import sys
import io
import requests
import json
from datetime import datetime
import openpyxl

# UTF-8 출력 설정
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

def get_access_token(app_key, app_secret):
    """OAuth 토큰 발급"""
    url = "https://openapi.koreainvestment.com:9443/oauth2/tokenP"
    headers = {"content-type": "application/json"}
    body = {
        "grant_type": "client_credentials",
        "appkey": app_key,
        "appsecret": app_secret
    }

    try:
        response = requests.post(url, headers=headers, data=json.dumps(body), timeout=10)
        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            print(f"[ERROR] Token failed: {response.status_code}")
            return None
    except Exception as e:
        print(f"[ERROR] Token error: {e}")
        return None

def get_usd_balance(app_key, app_secret, account_no, access_token, ticker="SOXL", price=1.0):
    """USD 예수금 조회 (매수가능금액조회 API - Phoenix Trading 실제 사용)"""
    # 계좌번호 파싱
    if '-' in account_no:
        cano, acnt_prdt_cd = account_no.split('-')
    else:
        cano = account_no[:8]
        acnt_prdt_cd = account_no[8:]

    # 올바른 API: 매수가능금액조회
    url = "https://openapi.koreainvestment.com:9443/uapi/overseas-stock/v1/trading/inquire-psamount"

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "TTTS3007R",  # 매수가능금액조회
        "custtype": "P"
    }

    params = {
        "CANO": cano,
        "ACNT_PRDT_CD": acnt_prdt_cd,
        "OVRS_EXCG_CD": "NASD",
        "OVRS_ORD_UNPR": f"{price:.2f}",
        "ITEM_CD": ticker
    }

    try:
        response = requests.get(url, headers=headers, params=params, timeout=10)

        if response.status_code == 200:
            data = response.json()
            rt_cd = data.get("rt_cd")

            if rt_cd == "0":
                output = data.get("output", {})
                # ord_psbl_frcr_amt: 주문가능외화금액 (USD 예수금)
                cash = float(output.get("ord_psbl_frcr_amt", 0))
                return cash, data
            elif rt_cd == "7":
                # 조회가능 상품이 없습니다 (정상)
                return 0.0, data
            else:
                print(f"[ERROR] Query failed: rt_cd={rt_cd}, msg={data.get('msg1')}")
                return None, data
        else:
            print(f"[ERROR] HTTP error: {response.status_code}")
            return None, None
    except Exception as e:
        print(f"[ERROR] Balance query error: {e}")
        return None, None

def main():
    print("=" * 60)
    print("Korea Investment USD Balance Check (Real)")
    print("=" * 60)
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Excel에서 설정 읽기
    excel_path = 'phoenix_grid_template_v3.xlsx'

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        app_key = ws['B12'].value
        app_secret = ws['B13'].value
        account_no = ws['B14'].value
        ticker = ws['B18'].value or "SOXL"

        print(f"[INFO] Reading from Excel: {excel_path}")
        print(f"Account: {account_no}")
        print(f"Ticker: {ticker}")
        print(f"APP KEY: {app_key[:20]}...") if app_key else print("APP KEY: None")
        print()

        if not app_key or not app_secret or not account_no:
            print("[ERROR] Excel settings incomplete!")
            print("Please check B12, B13, B14")
            return

    except FileNotFoundError:
        print(f"[ERROR] Excel file not found: {excel_path}")
        print("Run this in the 'release' folder.")
        return
    except Exception as e:
        print(f"[ERROR] Excel read failed: {e}")
        return

    # 토큰 발급
    print("Requesting OAuth token...")
    access_token = get_access_token(app_key, app_secret)

    if not access_token:
        print("[ERROR] Token issuance failed")
        return

    print("[OK] Token issued successfully")
    print()

    # USD 예수금 조회
    print(f"Querying USD balance for {ticker} (using TTTS3007R API)...")
    usd_balance, data = get_usd_balance(app_key, app_secret, account_no, access_token, ticker=ticker)

    print()
    print("=" * 60)

    if usd_balance is not None:
        print(f"[RESULT] USD Balance: ${usd_balance:,.2f}")

        if usd_balance == 0:
            print()
            print("[WARNING] USD balance is $0.00")
        else:
            print()
            print("[SUCCESS] Tradeable USD balance confirmed!")

        # API 응답 상세
        if data:
            print()
            print("-" * 60)
            print("API Response:")
            print(f"  rt_cd: {data.get('rt_cd')}")
            print(f"  msg1: {data.get('msg1', '').strip()}")
            print()
            print("Full API Response (DEBUG):")
            print(json.dumps(data, indent=2, ensure_ascii=False))
    else:
        print("[ERROR] Balance query failed")

    print("=" * 60)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nTerminated.")
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()

    print()
    input("Press Enter to exit...")
