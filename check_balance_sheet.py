"""
해외주식 잔고 시트 상세 분석
"""
import openpyxl
import sys

def analyze_balance_sheet(file_path):
    """해외주식 잔고 시트 상세 분석"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        print(f"\n{'='*80}")
        print(f"파일: {file_path}")
        print(f"{'='*80}\n")

        # 시트 목록 출력
        print(f"전체 시트 목록 ({len(wb.sheetnames)}개):")
        for i, name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {name}")

        # 해외주식 잔고 시트 찾기
        balance_sheet_name = None
        for name in wb.sheetnames:
            if '잔고' in name:
                balance_sheet_name = name
                break

        if not balance_sheet_name:
            print("\n[ERROR] '잔고' 시트를 찾을 수 없습니다.")
            return

        print(f"\n{'='*80}")
        print(f"[분석 대상] {balance_sheet_name}")
        print(f"{'='*80}\n")

        ws = wb[balance_sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        print(f"시트 크기: {max_row} 행 x {max_col} 열\n")

        # 모든 행 출력
        print(f"{'─'*80}")
        print("전체 데이터:")
        print(f"{'─'*80}\n")

        for row_idx in range(1, min(max_row + 1, 100)):  # 최대 100행까지
            row_data = []
            has_data = False

            for col_idx in range(1, max_col + 1):
                cell_value = ws.cell(row_idx, col_idx).value
                col_letter = openpyxl.utils.get_column_letter(col_idx)

                if cell_value is not None:
                    has_data = True
                    row_data.append(f"{col_letter}{row_idx}={cell_value}")

            if has_data:
                print(f"행 {row_idx:3d}: {' | '.join(row_data)}")

        print(f"\n{'─'*80}")
        print(f"총 {max_row} 행 중 데이터가 있는 행만 표시")
        print(f"{'─'*80}\n")

        # 특정 필드 검색 (USD, 잔고, 예수금 등)
        print(f"\n{'='*80}")
        print("[키워드 검색] USD, 잔고, 예수금, 현금")
        print(f"{'='*80}\n")

        keywords = ['USD', '잔고', '예수금', '현금', 'CASH', 'balance']
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and isinstance(cell_value, str):
                    for keyword in keywords:
                        if keyword.upper() in str(cell_value).upper():
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            print(f"발견: {col_letter}{row_idx} = {cell_value}")
                            break

    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    file_path = r"C:\Users\PC_1M\Downloads\[해외주식] 주문_계좌.xlsx"
    analyze_balance_sheet(file_path)
