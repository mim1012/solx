"""Excel 파일에서 API 키 확인"""
import sys
from openpyxl import load_workbook

excel_path = "release/phoenix_grid_template_v3.xlsx"

print("=" * 80)
print("Excel 파일 API 키 확인")
print("=" * 80)

try:
    wb = load_workbook(excel_path, data_only=False)
    ws = wb.active

    b12 = ws["B12"].value  # KIS_APP_KEY
    b13 = ws["B13"].value  # KIS_APP_SECRET
    b14 = ws["B14"].value  # KIS_ACCOUNT_NO
    b15 = ws["B15"].value  # SYSTEM_RUNNING

    print(f"\nExcel 파일: {excel_path}")
    print(f"\nB12 (KIS_APP_KEY): {b12}")
    if b12:
        print(f"  - 길이: {len(str(b12))} 자")
        print(f"  - 앞 4자: {str(b12)[:4]}****")

    print(f"\nB13 (KIS_APP_SECRET): {b13}")
    if b13:
        print(f"  - 길이: {len(str(b13))} 자")

    print(f"\nB14 (KIS_ACCOUNT_NO): {b14}")

    print(f"\nB15 (SYSTEM_RUNNING): {b15}")

    # 검증
    if not b12 or str(b12) == "your_app_key_here":
        print("\n❌ 오류: B12에 실제 KIS_APP_KEY를 입력하세요!")
        sys.exit(1)

    if not b13 or str(b13) == "your_app_secret_here":
        print("\n❌ 오류: B13에 실제 KIS_APP_SECRET을 입력하세요!")
        sys.exit(1)

    if not b14 or str(b14) == "12345678-01":
        print("\n❌ 오류: B14에 실제 계좌번호를 입력하세요!")
        sys.exit(1)

    print("\n✅ API 키가 모두 설정되어 있습니다!")

except Exception as e:
    print(f"\n❌ 오류: {e}")
    sys.exit(1)
