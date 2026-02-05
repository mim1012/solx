# -*- coding: utf-8 -*-
"""Check B22 cell in Excel template"""
import openpyxl

wb = openpyxl.load_workbook('phoenix_grid_template_v3.xlsx')
ws = wb['01_매매전략_기준설정']

print(f"A22: '{ws['A22'].value}'")
print(f"B22: '{ws['B22'].value}'")

# Check surrounding cells
for row in range(20, 25):
    a_val = ws[f'A{row}'].value
    b_val = ws[f'B{row}'].value
    print(f"A{row}: '{a_val}' | B{row}: '{b_val}'")

wb.close()
