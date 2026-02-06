"""
해외주식 주문_계좌 Excel 파일 분석 스크립트
"""
import openpyxl
import sys

def analyze_excel_file(file_path):
    """Excel 파일의 모든 시트와 데이터 분석"""
    try:
        # Excel 파일 열기
        wb = openpyxl.load_workbook(file_path, data_only=True)

        print(f"\n{'='*80}")
        print(f"파일: {file_path}")
        print(f"{'='*80}\n")

        print(f"총 시트 수: {len(wb.sheetnames)}")
        print(f"시트 목록: {wb.sheetnames}\n")

        # 각 시트 분석
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            print(f"\n{'─'*80}")
            print(f"[시트] {sheet_name}")
            print(f"{'─'*80}")

            # 시트 크기
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"크기: {max_row} 행 x {max_col} 열\n")

            # 헤더 행 출력 (첫 3행)
            print("[헤더] 초기 데이터 (첫 5행):")
            for row_idx in range(1, min(6, max_row + 1)):
                row_data = []
                for col_idx in range(1, min(max_col + 1, 20)):  # 최대 20열까지
                    cell_value = ws.cell(row_idx, col_idx).value
                    if cell_value is not None:
                        row_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}: {cell_value}")

                if row_data:
                    print(f"  행 {row_idx}: {' | '.join(row_data)}")

            # 데이터가 있는 행 확인
            print(f"\n[데이터 요약]")
            non_empty_rows = 0
            for row_idx in range(1, max_row + 1):
                row_has_data = any(ws.cell(row_idx, col_idx).value is not None
                                  for col_idx in range(1, max_col + 1))
                if row_has_data:
                    non_empty_rows += 1

            print(f"  - 데이터가 있는 행: {non_empty_rows}/{max_row}")

            # 마지막 몇 행 확인 (데이터가 있는 경우)
            if max_row > 5:
                print(f"\n[마지막 데이터] 최근 3행:")
                for row_idx in range(max(1, max_row - 2), max_row + 1):
                    row_data = []
                    for col_idx in range(1, min(max_col + 1, 20)):
                        cell_value = ws.cell(row_idx, col_idx).value
                        if cell_value is not None:
                            row_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}: {cell_value}")

                    if row_data:
                        print(f"  행 {row_idx}: {' | '.join(row_data)}")

        print(f"\n{'='*80}")
        print("분석 완료")
        print(f"{'='*80}\n")

    except Exception as e:
        print(f"[ERROR] 오류 발생: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    file_path = r"C:\Users\PC_1M\Downloads\[해외주식] 주문_계좌.xlsx"
    analyze_excel_file(file_path)
