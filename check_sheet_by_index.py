"""
시트 번호로 직접 조회
"""
import openpyxl

def analyze_sheet_by_index(file_path, sheet_index):
    """시트 번호로 직접 조회"""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    sheet_name = wb.sheetnames[sheet_index]

    print(f"\n{'='*80}")
    print(f"[시트 #{sheet_index+1}] {sheet_name}")
    print(f"{'='*80}\n")

    ws = wb[sheet_name]
    max_row = min(ws.max_row, 100)
    max_col = ws.max_column

    print(f"크기: {max_row} 행 x {max_col} 열\n")

    # 전체 데이터 출력
    for row_idx in range(1, max_row + 1):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                # 긴 값은 일부만 표시
                str_value = str(cell_value)
                if len(str_value) > 100:
                    str_value = str_value[:100] + "..."
                row_data.append(f"{col_letter}{row_idx}={str_value}")

        if row_data:
            print(f"행 {row_idx:3d}: {' | '.join(row_data)}")

    print(f"\n{'='*80}")

if __name__ == "__main__":
    file_path = r"C:\Users\PC_1M\Downloads\[해외주식] 주문_계좌.xlsx"

    # 6번, 10번, 12번 시트 확인
    for idx in [5, 9, 11]:  # 0-based index
        analyze_sheet_by_index(file_path, idx)
        print("\n\n")
