"""
한국투자증권 USD 예수금 조회 스크립트
"""
import openpyxl
import requests
import json
from datetime import datetime

def get_access_token(app_key, app_secret):
    """OAuth 토큰 발급"""
    url = "https://openapi.koreainvestment.com:9443/oauth2/tokenP"

    headers = {"content-type": "application/json"}
    body = {
        "grant_type": "client_credentials",
        "appkey": app_key,
        "appsecret": app_secret
    }

    response = requests.post(url, headers=headers, data=json.dumps(body))
    if response.status_code == 200:
        return response.json().get("access_token")
    else:
        print(f"토큰 발급 실패: {response.status_code}")
        print(response.text)
        return None

def get_usd_balance(app_key, app_secret, account_no, access_token):
    """USD 예수금 조회"""
    # 계좌번호 파싱
    if '-' in account_no:
        cano, acnt_prdt_cd = account_no.split('-')
    else:
        cano = account_no[:8]
        acnt_prdt_cd = account_no[8:]

    url = "https://openapi.koreainvestment.com:9443/uapi/overseas-stock/v1/trading/inquire-psamount"

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "TTTS3007R",  # 실전투자
        "custtype": "P"
    }

    params = {
        "CANO": cano,
        "ACNT_PRDT_CD": acnt_prdt_cd,
        "OVRS_EXCG_CD": "NASD",
        "OVRS_ORD_UNPR": "1.00",
        "ITEM_CD": "SOXL"
    }

    response = requests.get(url, headers=headers, params=params, timeout=10)

    if response.status_code == 200:
        data = response.json()
        rt_cd = data.get("rt_cd")

        if rt_cd == "0":
            output = data.get("output", {})
            usd_balance = float(output.get("ord_psbl_frcr_amt", 0))
            return usd_balance, data
        elif rt_cd == "7":
            return 0.0, data
        else:
            print(f"조회 실패: rt_cd={rt_cd}, msg={data.get('msg1')}")
            return None, data
    else:
        print(f"HTTP 오류: {response.status_code}")
        print(response.text)
        return None, None

def main():
    print("=" * 60)
    print("한국투자증권 USD 예수금 조회")
    print("=" * 60)
    print(f"조회 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Excel에서 설정 읽기
    try:
        wb = openpyxl.load_workbook('release/phoenix_grid_template_v3.xlsx')
        ws = wb.active

        app_key = ws['B12'].value
        app_secret = ws['B13'].value
        account_no = ws['B14'].value

        print(f"계좌번호: {account_no}")
        print(f"APP KEY: {app_key[:20]}..." if app_key else "APP KEY: None")
        print()

        if not app_key or not app_secret or not account_no:
            print("[ERROR] Excel 파일에 API 키 또는 계좌번호가 설정되지 않았습니다.")
            print("   B12: KIS APP KEY")
            print("   B13: KIS APP SECRET")
            print("   B14: 계좌번호 (예: 12345678-01)")
            return

    except Exception as e:
        print(f"[ERROR] Excel 파일 읽기 실패: {e}")
        return

    # 토큰 발급
    print("토큰 발급 중...")
    access_token = get_access_token(app_key, app_secret)

    if not access_token:
        print("[ERROR] OAuth 토큰 발급 실패")
        return

    print("[OK] 토큰 발급 성공")
    print()

    # USD 예수금 조회
    print("USD 예수금 조회 중...")
    usd_balance, data = get_usd_balance(app_key, app_secret, account_no, access_token)

    print()
    print("=" * 60)

    if usd_balance is not None:
        print(f"[OK] USD 예수금: ${usd_balance:,.2f}")

        if usd_balance == 0:
            print()
            print("[WARNING] USD 예수금이 0원입니다.")
            print("   - 원화를 USD로 환전하지 않았거나")
            print("   - 해외주식 거래 권한이 없을 수 있습니다.")
            print()
            print("해결 방법:")
            print("   1. 한국투자증권 HTS 로그인")
            print("   2. 원화 -> USD 환전")
            print("   3. 최소 600달러 이상 권장")

        # 상세 응답 출력
        if data:
            print()
            print("-" * 60)
            print("API 응답 상세:")
            print(f"  rt_cd: {data.get('rt_cd')}")
            print(f"  msg1: {data.get('msg1')}")
            if data.get('output'):
                output = data.get('output')
                print(f"  ord_psbl_frcr_amt (주문가능금액): ${output.get('ord_psbl_frcr_amt', 'N/A')}")
    else:
        print("[ERROR] USD 예수금 조회 실패")

    print("=" * 60)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n사용자가 중단했습니다.")
    except Exception as e:
        print(f"\n예외 발생: {e}")
        import traceback
        traceback.print_exc()
