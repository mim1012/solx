"""
한국투자증권 USD 예수금 조회 (Excel 자동 읽기)
"""
import sys
import io
import requests
import json
from datetime import datetime
import openpyxl

# UTF-8 출력 설정 (한글 깨짐 방지)
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

def get_access_token(app_key, app_secret):
    """OAuth 토큰 발급"""
    url = "https://openapi.koreainvestment.com:9443/oauth2/tokenP"

    headers = {"content-type": "application/json"}
    body = {
        "grant_type": "client_credentials",
        "appkey": app_key,
        "appsecret": app_secret
    }

    try:
        response = requests.post(url, headers=headers, data=json.dumps(body), timeout=10)
        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            print(f"[ERROR] Token issuance failed: {response.status_code}")
            print(response.text)
            return None
    except Exception as e:
        print(f"[ERROR] Token request error: {e}")
        return None

def get_usd_balance(app_key, app_secret, account_no, access_token):
    """USD 예수금 조회"""
    # 계좌번호 파싱
    if '-' in account_no:
        cano, acnt_prdt_cd = account_no.split('-')
    else:
        cano = account_no[:8]
        acnt_prdt_cd = account_no[8:]

    url = "https://openapi.koreainvestment.com:9443/uapi/overseas-stock/v1/trading/inquire-psamount"

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "TTTS3007R",  # 실전투자
        "custtype": "P"
    }

    params = {
        "CANO": cano,
        "ACNT_PRDT_CD": acnt_prdt_cd,
        "OVRS_EXCG_CD": "NASD",
        "OVRS_ORD_UNPR": "1.00",
        "ITEM_CD": "SOXL"
    }

    try:
        response = requests.get(url, headers=headers, params=params, timeout=10)

        if response.status_code == 200:
            data = response.json()
            rt_cd = data.get("rt_cd")

            if rt_cd == "0":
                output = data.get("output", {})
                usd_balance = float(output.get("ord_psbl_frcr_amt", 0))
                return usd_balance, data
            elif rt_cd == "7":
                return 0.0, data
            else:
                print(f"[ERROR] Query failed: rt_cd={rt_cd}, msg={data.get('msg1')}")
                return None, data
        else:
            print(f"[ERROR] HTTP error: {response.status_code}")
            print(response.text)
            return None, None
    except Exception as e:
        print(f"[ERROR] Balance query error: {e}")
        return None, None

def main():
    print("=" * 60)
    print("Korea Investment USD Balance Check")
    print("=" * 60)
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Excel에서 설정 읽기
    excel_path = 'phoenix_grid_template_v3.xlsx'

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        app_key = ws['B12'].value
        app_secret = ws['B13'].value
        account_no = ws['B14'].value

        print(f"[INFO] Reading from Excel: {excel_path}")
        print(f"Account: {account_no}")
        print(f"APP KEY: {app_key[:20]}..." if app_key else "APP KEY: None")
        print()

        if not app_key or not app_secret or not account_no:
            print("[ERROR] Excel settings incomplete!")
            print()
            print("Please check:")
            print("  B12: KIS APP KEY")
            print("  B13: KIS APP SECRET")
            print("  B14: Account Number (e.g., 12345678-01)")
            return

    except FileNotFoundError:
        print(f"[ERROR] Excel file not found: {excel_path}")
        print()
        print("Make sure to run this program in the 'release' folder")
        print("where phoenix_grid_template_v3.xlsx is located.")
        return
    except Exception as e:
        print(f"[ERROR] Failed to read Excel: {e}")
        return

    # 토큰 발급
    print("Requesting OAuth token...")
    access_token = get_access_token(app_key, app_secret)

    if not access_token:
        print("[ERROR] OAuth token issuance failed")
        print()
        print("Check:")
        print("  1. APP KEY and APP SECRET are correct")
        print("  2. Using real trading APP KEY (not mock)")
        print("  3. APP status at https://apiportal.koreainvestment.com")
        return

    print("[OK] Token issued successfully")
    print()

    # USD 예수금 조회
    print("Querying USD balance...")
    usd_balance, data = get_usd_balance(app_key, app_secret, account_no, access_token)

    print()
    print("=" * 60)

    if usd_balance is not None:
        print(f"[RESULT] USD Balance: ${usd_balance:,.2f}")

        if usd_balance == 0:
            print()
            print("[WARNING] USD balance is $0.00")
            print()
            print("Possible reasons:")
            print("  1. KRW not converted to USD")
            print("  2. No overseas trading authorization")
            print("  3. Actually zero balance")
            print()
            print("Solution:")
            print("  1. Login to Korea Investment HTS")
            print("  2. Go to [Overseas Stock] -> [Exchange]")
            print("  3. Exchange KRW -> USD (min. $600 recommended)")
            print()
            print("Note:")
            print("  - USD balance required for Phoenix Trading")
            print("  - Check again after 5-10 minutes post-exchange")
        else:
            print()
            print("[SUCCESS] Tradeable USD balance confirmed!")
            print("You can run Phoenix Trading now.")

        # 상세 응답
        if data:
            print()
            print("-" * 60)
            print("API Response Detail:")
            print(f"  rt_cd: {data.get('rt_cd')}")
            print(f"  msg1: {data.get('msg1', '').strip()}")
            if data.get('output'):
                output = data.get('output')
                print(f"  ord_psbl_frcr_amt: ${output.get('ord_psbl_frcr_amt', 'N/A')}")
    else:
        print("[ERROR] USD balance query failed")
        print()
        print("Please verify your account number.")

    print("=" * 60)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nProgram terminated.")
    except Exception as e:
        print(f"\nError occurred: {e}")
        import traceback
        traceback.print_exc()

    print()
    input("Press Enter to exit...")
