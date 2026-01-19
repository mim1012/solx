"""
Phoenix Grid System - Excel 템플릿 생성기 v3.1 (커스텀 버전)
거래 종목: SOXL (실거래 전용)

원본 Run.exe Excel 구조 기반 + 커스텀 기능 추가:
- 시트 1: 01_매매전략_기준설정 (마스터/컨트롤)
- 시트 2: 02_운용로그_히스토리 (시간 순 기록)

커스텀 v3.1 변경사항:
- Tier 1 매수/매도 기능 추가 (사용자 선택 가능)
- "1티어 거래 여부" 설정 추가 (TRUE/FALSE)
- Tier 1 매수% 사용자 직접 설정 가능
- Excel 업데이트 주기 설정 추가 (기본 1초)
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_phoenix_template():
    """Phoenix 그리드 시스템용 Excel 템플릿 생성 (원본 구조)"""
    wb = openpyxl.Workbook()

    # 기본 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    warning_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==================== 시트 1: 01_매매전략_기준설정 ====================
    ws_master = wb.active
    ws_master.title = "01_매매전략_기준설정"

    # ===== A. 기본 설정 영역 (A1:B17) =====
    settings_data = [
        ["=== 기본 설정 ===", ""],
        ["계좌번호", ""],
        ["종목코드", "SOXL"],
        ["투자금 (USD)", "10000"],
        ["티어 분할 수", "240"],
        ["1티어 금액 (USD)", "500"],
        ["1티어 갱신 여부", "TRUE"],
        ["1티어 거래 여부", "FALSE"],  # 신규: Tier 1 매수/매도 활성화
        ["매수 제한", "FALSE"],
        ["매도 제한", "FALSE"],
        ["", ""],
        ["=== 텔레그램 알림 ===", ""],
        ["텔레그램 ID", ""],
        ["텔레그램 Token", ""],
        ["알림 활성화", "FALSE"],
        ["Excel 업데이트 주기 (초)", "40.0"],  # 신규: Excel 저장 주기 설정 (기본 40초)
    ]

    for row_idx, row_data in enumerate(settings_data, start=1):
        ws_master.cell(row=row_idx, column=1, value=row_data[0])
        ws_master.cell(row=row_idx, column=2, value=row_data[1])

        # 섹션 헤더 스타일
        if "===" in str(row_data[0]):
            cell = ws_master.cell(row=row_idx, column=1)
            cell.fill = header_fill
            cell.font = header_font
            ws_master.merge_cells(f'A{row_idx}:B{row_idx}')

    ws_master.column_dimensions['A'].width = 25
    ws_master.column_dimensions['B'].width = 20

    # ===== B. 프로그램 매매 정보 (D1:E10) =====
    program_info_data = [
        ["=== 프로그램 매매 정보 ===", ""],
        ["최근 업데이트 시간", ""],
        ["현재 티어", "1"],
        ["현재가 (USD)", "0.00"],
        ["잔고 (USD)", "0.00"],
        ["수량차 (주)", "0"],
        ["현재 매수 상태", "대기"],
        ["현재 매도 상태", "대기"],
    ]

    for row_idx, row_data in enumerate(program_info_data, start=1):
        ws_master.cell(row=row_idx, column=4, value=row_data[0])
        ws_master.cell(row=row_idx, column=5, value=row_data[1])

        if "===" in str(row_data[0]):
            cell = ws_master.cell(row=row_idx, column=4)
            cell.fill = header_fill
            cell.font = header_font
            ws_master.merge_cells(f'D{row_idx}:E{row_idx}')

    ws_master.column_dimensions['D'].width = 25
    ws_master.column_dimensions['E'].width = 20

    # ===== C. 매매 기준 설정 테이블 (A17:E257, 240개 행) =====
    table_start_row = 17

    # 헤더
    table_headers = ["티어", "시드비%", "매수%", "매도%", "비고"]
    for col_idx, header in enumerate(table_headers, start=1):
        cell = ws_master.cell(row=table_start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    # 240개 티어 데이터
    for tier in range(1, 241):
        row_idx = table_start_row + tier
        seed_pct = 0.5 * (tier - 1)  # 0%, 0.5%, 1.0%, ...
        buy_pct = -seed_pct if tier > 1 else 0
        sell_pct = 3.0
        remark = "기준가 (거래 설정 시 매매 가능)" if tier == 1 else ""

        ws_master.cell(row=row_idx, column=1, value=tier)
        ws_master.cell(row=row_idx, column=2, value=seed_pct / 100).number_format = '0.0%'
        # Tier 1: 기본 0.0%, 사용자가 "1티어 거래 여부=TRUE"로 설정 시 수정 가능
        ws_master.cell(row=row_idx, column=3, value=buy_pct / 100).number_format = '0.0%'
        ws_master.cell(row=row_idx, column=4, value=sell_pct / 100).number_format = '0.0%'
        ws_master.cell(row=row_idx, column=5, value=remark)

        for col_idx in range(1, 6):
            ws_master.cell(row=row_idx, column=col_idx).border = border_thin

    # ===== D. 프로그램 영역 (시뮬레이션, G17:N257) =====
    sim_headers = ["티어", "잔고량", "투자금", "티어평단", "매수(가)", "매수(량)", "매도(가)", "매도(량)"]
    for col_idx, header in enumerate(sim_headers, start=7):
        cell = ws_master.cell(row=table_start_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    # 240개 티어 시뮬레이션 데이터
    for tier in range(1, 241):
        row_idx = table_start_row + tier

        ws_master.cell(row=row_idx, column=7, value=tier)  # 티어
        ws_master.cell(row=row_idx, column=8, value=0)  # 잔고량
        ws_master.cell(row=row_idx, column=9, value=0).number_format = '$#,##0.00'  # 투자금
        ws_master.cell(row=row_idx, column=10, value=0).number_format = '$#,##0.00'  # 티어평단
        ws_master.cell(row=row_idx, column=11, value="").number_format = '$#,##0.00'  # 매수(가)
        ws_master.cell(row=row_idx, column=12, value="")  # 매수(량)
        ws_master.cell(row=row_idx, column=13, value="").number_format = '$#,##0.00'  # 매도(가)
        ws_master.cell(row=row_idx, column=14, value="")  # 매도(량)

        for col_idx in range(7, 15):
            ws_master.cell(row=row_idx, column=col_idx).border = border_thin

    # 컬럼 너비 조정
    for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws_master.column_dimensions[col].width = 12

    # ==================== 시트 2: 02_운용로그_히스토리 ====================
    ws_history = wb.create_sheet("02_운용로그_히스토리")

    # 헤더
    history_headers = [
        # A. 시간·기본 정보
        "업데이트", "날짜", "시트", "종목", "티어",
        # B. 포지션 상태
        "총티어", "잔고량(차)", "투자금", "1티어",
        # C. 손익·성과 지표
        "예수금", "주식평가금", "잔고수익", "매수예정", "인출가능", "아비타수익",
        # D. 실제 체결 수량
        "매수", "매도"
    ]

    for col_idx, header in enumerate(history_headers, start=1):
        cell = ws_history.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    # 샘플 데이터 (10개 행)
    for row_idx in range(2, 12):
        ws_history.cell(row=row_idx, column=1, value="").number_format = 'YYYY-MM-DD HH:MM:SS'  # 업데이트
        ws_history.cell(row=row_idx, column=2, value="").number_format = 'YYYY-MM-DD'  # 날짜
        ws_history.cell(row=row_idx, column=3, value="Main")  # 시트
        ws_history.cell(row=row_idx, column=4, value="SOXL")  # 종목
        for col_idx in range(5, 18):
            cell = ws_history.cell(row=row_idx, column=col_idx, value="")
            cell.border = border_thin

            # 숫자 포맷
            if col_idx in [8, 9, 10, 11, 13, 14]:  # 금액 컬럼
                cell.number_format = '$#,##0.00'
            elif col_idx in [12, 15]:  # 수익률 컬럼
                cell.number_format = '0.00%'

    # 컬럼 너비 조정
    ws_history.column_dimensions['A'].width = 20  # 업데이트
    ws_history.column_dimensions['B'].width = 12  # 날짜
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws_history.column_dimensions[col_letter].width = 12

    # 파일 저장
    output_path = "D:\\Project\\SOLX\\phoenix_grid_template_v3.xlsx"
    wb.save(output_path)
    print(f"[OK] Excel 템플릿 생성 완료 (커스텀 v3.1): {output_path}")
    print(f"[INFO] 시트 1: 01_매매전략_기준설정 (240개 티어 테이블)")
    print(f"[INFO] 시트 2: 02_운용로그_히스토리 (시간 순 누적)")
    print(f"[INFO] 종목: SOXL (3X 레버리지 ETF)")
    print(f"[WARN] 실거래 전용 - 모의투자 미지원")
    print(f"[NEW] 텔레그램 알림 필드 추가")
    print(f"[NEW] 매수/매도 제한 스위치 추가")
    print(f"[NEW] 1티어 갱신 여부 제어 추가 (TRUE=자동상승, FALSE=고정)")
    print(f"[CUSTOM v3.1] 1티어 거래 여부 추가 (기본값: FALSE, TRUE 시 Tier 1 매수/매도 가능)")
    print(f"[CUSTOM v3.1] Tier 1 매수% 사용자 수동 설정 가능 (기본값: 0.0%)")
    print(f"[INFO] 프로그램 영역 컬럼: 티어, 잔고량, 투자금, 티어평단, 매수(가), 매수(량), 매도(가), 매도(량)")
    return output_path

if __name__ == "__main__":
    create_phoenix_template()
