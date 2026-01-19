"""
Phoenix 자동매매 시스템 데이터 무결성 검증 스크립트

Excel 파일의 데이터 일관성을 체크합니다:
- 시트 1 "프로그램 영역" 보유량 vs 시트 2 "운용로그" 계산 보유량
- 매수/매도 기록 누락 확인
- 시간 순서 검증
"""

import openpyxl
import pandas as pd
from datetime import datetime
from typing import Tuple, List, Dict


def verify_data_integrity(excel_path: str) -> Tuple[bool, List[str]]:
    """
    Excel 파일의 데이터 무결성 검증

    Args:
        excel_path: Excel 파일 경로

    Returns:
        (bool, List[str]): (검증 성공 여부, 오류/경고 메시지 리스트)
    """
    errors = []
    warnings = []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except FileNotFoundError:
        return False, [f"❌ 파일을 찾을 수 없습니다: {excel_path}"]
    except Exception as e:
        return False, [f"❌ 파일 로드 실패: {e}"]

    # ===========================
    # 1. 시트 1: 프로그램 영역 검증
    # ===========================
    print("\n📊 시트 1: 프로그램 영역 검증")

    try:
        ws1 = wb["01_매매전략_기준설정"]
    except KeyError:
        return False, [f"❌ 시트를 찾을 수 없습니다: 01_매매전략_기준설정"]

    # 총 보유량 계산 (H열: 잔고량)
    total_qty_sheet1 = 0
    invested_amount_sheet1 = 0
    active_tiers = []

    for tier in range(1, 241):
        row_idx = 17 + tier  # 헤더는 17행, 데이터는 18행부터

        qty = ws1[f"H{row_idx}"].value or 0
        invested = ws1[f"I{row_idx}"].value or 0

        if qty > 0:
            total_qty_sheet1 += qty
            invested_amount_sheet1 += invested
            active_tiers.append({
                "tier": tier,
                "qty": qty,
                "invested": invested,
                "avg_price": ws1[f"J{row_idx}"].value or 0
            })

    print(f"  총 보유량: {total_qty_sheet1}주")
    print(f"  투자금: ${invested_amount_sheet1:.2f}")
    print(f"  활성 티어 수: {len(active_tiers)}개")

    # ===========================
    # 2. 시트 2: 운용로그 검증
    # ===========================
    print("\n📝 시트 2: 운용로그 검증")

    try:
        ws2 = wb["02_운용로그_히스토리"]
    except KeyError:
        return False, [f"❌ 시트를 찾을 수 없습니다: 02_운용로그_히스토리"]

    # 데이터 읽기
    data = []
    for row in ws2.iter_rows(min_row=2, values_only=True):  # 헤더 제외
        if row[0]:  # 업데이트 시간이 있는 행만
            data.append(row)

    if not data:
        warnings.append("⚠️ 운용로그에 기록이 없습니다 (시스템 시작 직후일 수 있음)")
        print("  기록 없음")
    else:
        # DataFrame 생성
        df = pd.DataFrame(data, columns=[
            "업데이트", "날짜", "시트", "종목", "티어", "총티어", "잔고량(차)",
            "투자금", "1티어", "예수금", "주식평가금", "잔고수익",
            "매수예정", "인출가능", "아비타수익", "매수", "매도"
        ])

        # 매수/매도 합계 계산
        total_buy = df["매수"].fillna(0).sum()
        total_sell = df["매도"].fillna(0).sum()
        expected_qty = total_buy - total_sell

        print(f"  총 기록 수: {len(df)}개")
        print(f"  총 매수: {total_buy}주")
        print(f"  총 매도: {total_sell}주")
        print(f"  계산 보유량: {expected_qty}주")

        # 최근 기록의 보유량
        latest_qty = df["잔고량(차)"].iloc[-1]
        print(f"  최근 기록 보유량: {latest_qty}주")

        # ===========================
        # 3. 데이터 일치 검증
        # ===========================
        print("\n✅ 데이터 일치 검증")

        # 보유량 일치 검증
        if total_qty_sheet1 == expected_qty:
            print(f"  ✅ 보유량 일치: {total_qty_sheet1}주 = {expected_qty}주")
        else:
            error_msg = f"❌ 보유량 불일치! 프로그램 영역: {total_qty_sheet1}주, 운용로그 계산: {expected_qty}주"
            errors.append(error_msg)
            print(f"  {error_msg}")

        # 최근 기록 보유량 일치 검증
        if total_qty_sheet1 == latest_qty:
            print(f"  ✅ 최근 기록 일치: {total_qty_sheet1}주 = {latest_qty}주")
        else:
            warning_msg = f"⚠️ 최근 기록 불일치 (업데이트 지연 가능): 프로그램 영역: {total_qty_sheet1}주, 최근 기록: {latest_qty}주"
            warnings.append(warning_msg)
            print(f"  {warning_msg}")

        # ===========================
        # 4. 시간 순서 검증
        # ===========================
        print("\n🕐 시간 순서 검증")

        timestamps = []
        for idx, row in df.iterrows():
            try:
                ts = pd.to_datetime(row["업데이트"])
                timestamps.append(ts)
            except:
                warning_msg = f"⚠️ 행 {idx + 2}: 잘못된 시간 형식"
                warnings.append(warning_msg)

        if timestamps:
            sorted_timestamps = sorted(timestamps)
            if timestamps == sorted_timestamps:
                print(f"  ✅ 시간 순서 정상 (총 {len(timestamps)}개 기록)")
            else:
                error_msg = f"❌ 시간 순서 오류 발견! 정렬되지 않은 기록 있음"
                errors.append(error_msg)
                print(f"  {error_msg}")

        # ===========================
        # 5. 거래 기록 검증
        # ===========================
        print("\n💰 거래 기록 검증")

        buy_records = df[df["매수"] > 0]
        sell_records = df[df["매도"] > 0]

        print(f"  총 매수 거래: {len(buy_records)}건")
        print(f"  총 매도 거래: {len(sell_records)}건")

        # 모든 거래에 주문번호가 있는지 확인 (향후 추가 시)
        # 현재는 생략

        # ===========================
        # 6. 활성 티어 검증
        # ===========================
        print("\n🎯 활성 티어 검증")

        if active_tiers:
            print(f"  활성 티어 목록:")
            for pos in active_tiers[:10]:  # 최대 10개만 출력
                print(f"    - Tier {pos['tier']}: {pos['qty']}주 @ ${pos['avg_price']:.2f}")

            if len(active_tiers) > 10:
                print(f"    ... 외 {len(active_tiers) - 10}개 티어")
        else:
            print(f"  보유 포지션 없음")

    # ===========================
    # 7. 최종 결과
    # ===========================
    print("\n" + "=" * 60)

    if errors:
        print("❌ 데이터 무결성 검증 실패!")
        for error in errors:
            print(f"  {error}")
        print("\n⚠️ 경고:")
        for warning in warnings:
            print(f"  {warning}")
        return False, errors + warnings

    elif warnings:
        print("⚠️ 데이터 무결성 검증 완료 (경고 있음)")
        for warning in warnings:
            print(f"  {warning}")
        return True, warnings

    else:
        print("✅ 데이터 무결성 검증 성공!")
        print("  모든 데이터가 일치하며 이상 없음")
        return True, []


def generate_summary_report(excel_path: str) -> Dict:
    """
    요약 보고서 생성

    Args:
        excel_path: Excel 파일 경로

    Returns:
        dict: 요약 통계
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except:
        return {}

    # 시트 1: 기본 설정 읽기
    ws1 = wb["01_매매전략_기준설정"]

    settings = {
        "account_no": ws1["B2"].value,
        "ticker": ws1["B3"].value,
        "investment_usd": ws1["B4"].value,
        "tier1_price": ws1["E5"].value,  # 프로그램 매매 정보 - 1티어
        "current_price": ws1["E4"].value,  # 현재가
        "total_qty": ws1["E6"].value,  # 수량차
    }

    # 시트 2: 거래 통계
    ws2 = wb["02_운용로그_히스토리"]

    data = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data.append(row)

    if data:
        df = pd.DataFrame(data, columns=[
            "업데이트", "날짜", "시트", "종목", "티어", "총티어", "잔고량(차)",
            "투자금", "1티어", "예수금", "주식평가금", "잔고수익",
            "매수예정", "인출가능", "아비타수익", "매수", "매도"
        ])

        total_buy_count = len(df[df["매수"] > 0])
        total_sell_count = len(df[df["매도"] > 0])
        latest_record = df.iloc[-1]

        statistics = {
            "total_records": len(df),
            "total_buy_count": total_buy_count,
            "total_sell_count": total_sell_count,
            "realized_pnl_pct": latest_record["아비타수익"],
            "unrealized_pnl_pct": latest_record["잔고수익"],
        }
    else:
        statistics = {
            "total_records": 0,
            "total_buy_count": 0,
            "total_sell_count": 0,
            "realized_pnl_pct": 0,
            "unrealized_pnl_pct": 0,
        }

    return {
        "settings": settings,
        "statistics": statistics,
        "generated_at": datetime.now().isoformat()
    }


def main():
    """
    메인 실행 함수
    """
    import sys

    # Excel 파일 경로
    excel_path = "phoenix_grid_template_v3.xlsx"

    if len(sys.argv) > 1:
        excel_path = sys.argv[1]

    print("=" * 60)
    print("Phoenix 자동매매 시스템 - 데이터 무결성 검증")
    print("=" * 60)
    print(f"\n📁 파일: {excel_path}")
    print(f"🕐 실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 무결성 검증 실행
    success, messages = verify_data_integrity(excel_path)

    # 요약 보고서
    print("\n" + "=" * 60)
    print("📊 요약 보고서")
    print("=" * 60)

    summary = generate_summary_report(excel_path)

    if summary:
        settings = summary.get("settings", {})
        statistics = summary.get("statistics", {})

        print(f"\n🔧 설정:")
        print(f"  종목: {settings.get('ticker', 'N/A')}")
        print(f"  투자금: ${settings.get('investment_usd', 0):.2f}")
        print(f"  Tier 1: ${settings.get('tier1_price', 0):.2f}")
        print(f"  현재가: ${settings.get('current_price', 0):.2f}")
        print(f"  보유량: {settings.get('total_qty', 0)}주")

        print(f"\n📈 거래 통계:")
        print(f"  총 기록: {statistics.get('total_records', 0)}개")
        print(f"  매수 거래: {statistics.get('total_buy_count', 0)}건")
        print(f"  매도 거래: {statistics.get('total_sell_count', 0)}건")
        print(f"  실현 수익률: {statistics.get('realized_pnl_pct', 0):.2f}%")
        print(f"  미실현 수익률: {statistics.get('unrealized_pnl_pct', 0):.2f}%")

    print("\n" + "=" * 60)

    if success:
        print("✅ 검증 완료: 데이터 무결성 정상")
        sys.exit(0)
    else:
        print("❌ 검증 실패: 데이터 불일치 또는 오류 발견")
        sys.exit(1)


if __name__ == "__main__":
    main()
