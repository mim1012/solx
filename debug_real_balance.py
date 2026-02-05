"""
실제 USD 잔고 조회 디버그 스크립트
"""
import os
import sys
import openpyxl
import json
from pathlib import Path

# src 모듈 import
sys.path.insert(0, str(Path(__file__).parent))
from src.kis_rest_adapter import KisRestAdapter

def main():
    print("\n" + "="*80)
    print("USD 잔고 조회 디버그 - 실제 API 응답 확인")
    print("="*80 + "\n")

    # 1. Excel에서 설정 읽기
    excel_path = "phoenix_grid_template_v3.xlsx"

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)  # 수식도 읽기
        ws = wb["01_매매전략_기준설정"]

        # 실제 값 vs 수식 확인
        print("[Excel 설정 확인]")
        print(f"B12 (APP KEY) 값: {ws['B12'].value}")
        print(f"B12 (APP KEY) 수식: {ws['B12'].data_type}")
        print(f"B13 (APP SECRET) 값: {ws['B13'].value}")
        print(f"B14 (계좌번호): {ws['B14'].value}")
        print(f"B16 (티커): {ws['B16'].value}")
        print()

        # .env 파일에서도 읽어보기
        from dotenv import load_dotenv
        load_dotenv()

        app_key = os.getenv("KIS_APP_KEY")
        app_secret = os.getenv("KIS_APP_SECRET")
        account_no = os.getenv("KIS_ACCOUNT_NO")

        print("[.env 파일 확인]")
        print(f"KIS_APP_KEY: {app_key[:10]}... (앞 10자)" if app_key else "KIS_APP_KEY: None")
        print(f"KIS_APP_SECRET: {app_secret[:10]}... (앞 10자)" if app_secret else "KIS_APP_SECRET: None")
        print(f"KIS_ACCOUNT_NO: {account_no}")
        print()

        # 우선순위: .env > Excel
        if not app_key:
            app_key = ws['B12'].value
        if not app_secret:
            app_secret = ws['B13'].value
        if not account_no:
            account_no = ws['B14'].value

        ticker = ws['B16'].value

        if not app_key or not app_secret or not account_no:
            print("[ERROR] API 키 또는 계좌번호가 없습니다!")
            return

        print(f"[최종 사용 설정]")
        print(f"APP KEY: {str(app_key)[:10]}...")
        print(f"계좌번호: {account_no}")
        print(f"티커: {ticker}")
        print()

    except Exception as e:
        print(f"[ERROR] Excel 읽기 실패: {e}")
        return

    # 2. KIS API 어댑터 생성
    try:
        print("[KIS API 인증 시작]")
        kis = KisRestAdapter(
            app_key=str(app_key),
            app_secret=str(app_secret),
            account_no=str(account_no)
        )

        # 로그인
        if not kis.login():
            print("[ERROR] KIS API 로그인 실패!")
            return

        print(f"✅ 로그인 성공")
        print(f"Access Token: {kis.access_token[:20]}...")
        print()

    except Exception as e:
        print(f"[ERROR] KIS API 초기화 실패: {e}")
        import traceback
        traceback.print_exc()
        return

    # 3. 현재가 조회
    print("[SOXL 현재가 조회]")
    try:
        price = kis.get_overseas_price("SOXL")  # [FIX] 자동 거래소 감지 사용
        print(f"SOXL 현재가: ${price}")
        print()
    except Exception as e:
        print(f"현재가 조회 실패: {e}")
        price = 50.0  # 기본값
        print(f"기본값 사용: ${price}")
        print()

    # 4. USD 예수금 조회 (여러 거래소 코드로 시도)
    print("="*80)
    print("[USD 예수금 조회 테스트]")
    print("="*80 + "\n")

    exchange_codes = ["AMEX", "NASD", "NYSE"]  # [FIX] AMEX 우선, AMS 제거

    for exchange_code in exchange_codes:
        print(f"\n{'─'*80}")
        print(f"테스트 {exchange_code}: SOXL @ ${price}")
        print(f"{'─'*80}")

        try:
            # get_cash_balance 직접 호출
            import requests

            cano = account_no.split('-')[0]
            acnt_prdt_cd = account_no.split('-')[1]

            url = f"{kis.BASE_URL}/uapi/overseas-stock/v1/trading/inquire-psamount"

            params = {
                "CANO": cano,
                "ACNT_PRDT_CD": acnt_prdt_cd,
                "OVRS_EXCG_CD": exchange_code,
                "OVRS_ORD_UNPR": str(price),
                "ITEM_CD": "SOXL"
            }

            headers = kis._get_headers(tr_id="TTTS3007R")

            print(f"요청 URL: {url}")
            print(f"요청 파라미터: {json.dumps(params, indent=2)}")
            print(f"TR_ID: TTTS3007R")
            print()

            response = requests.get(url, headers=headers, params=params, timeout=10)

            print(f"HTTP Status: {response.status_code}")

            if response.status_code == 200:
                data = response.json()

                print(f"rt_cd: {data.get('rt_cd')}")
                print(f"msg_cd: {data.get('msg_cd')}")
                print(f"msg1: {data.get('msg1')}")
                print()

                if data.get("rt_cd") == "0":
                    output = data.get("output", {})

                    print("[Response Output]")
                    print(json.dumps(output, indent=2, ensure_ascii=False))
                    print()

                    # 예수금 추출
                    cash = float(output.get("ord_psbl_frcr_amt", 0) or 0)
                    print(f"✅ {exchange_code} 성공: USD 예수금 = ${cash:.2f}")

                    if cash > 0:
                        print(f"\n🎯 정답 발견! 거래소 코드: {exchange_code}")
                        print(f"USD 예수금: ${cash:.2f}")
                        break

                elif data.get("rt_cd") == "7":
                    print(f"⚠️ {exchange_code}: 상품이 없습니다 (rt_cd=7)")
                    print(f"   → 이 거래소에서는 SOXL 거래 이력이 없거나 잔고 없음")
                else:
                    print(f"❌ {exchange_code} 실패: {data.get('msg1')}")
                    print(f"전체 응답: {json.dumps(data, indent=2, ensure_ascii=False)}")
            else:
                print(f"❌ HTTP 오류: {response.status_code}")
                print(f"응답: {response.text}")

        except Exception as e:
            print(f"❌ {exchange_code} 예외: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "="*80)
    print("디버그 완료")
    print("="*80 + "\n")

if __name__ == "__main__":
    main()
