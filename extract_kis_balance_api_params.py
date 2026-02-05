"""
KIS API 잔고조회 파라미터 상세 추출
"""

import openpyxl
from pathlib import Path

KIS_DOC_PATH = r"C:\Users\PC_1M\Downloads\[해외주식] 주문_계좌.xlsx"


def extract_balance_api_params():
    """해외주식 체결기준현재잔고 API 파라미터 추출"""

    wb = openpyxl.load_workbook(KIS_DOC_PATH, data_only=True)

    print("=" * 80)
    print("[1] 해외주식 체결기준현재잔고 (현재 사용 중)")
    print("=" * 80)

    ws = wb["해외주식 체결기준현재잔고"]

    # Request Parameters 섹션 찾기
    print("\n[Request Parameters]")
    in_request_params = False
    param_rows = []

    for row in ws.iter_rows(values_only=True):
        # Request Parameters 섹션 시작
        if row[0] and "Request Parameters" in str(row[0]):
            in_request_params = True
            continue

        # Response 섹션 시작하면 종료
        if in_request_params and row[0] and "Response" in str(row[0]):
            break

        # 파라미터 행 수집
        if in_request_params and row[0]:
            param_rows.append(row[:7])  # 첫 7개 컬럼만

    # 파라미터 출력
    for row in param_rows[:20]:  # 최대 20개
        if row[0]:  # Element 필드가 있는 경우만
            element = row[0] or ""
            korean = row[1] or ""
            type_val = row[2] or ""
            required = row[3] or ""
            length = row[4] or ""
            desc = row[5] or ""

            print(f"\n  - {element}")
            print(f"    한글명: {korean}")
            print(f"    타입: {type_val}")
            print(f"    필수: {required}")
            print(f"    길이: {length}")
            print(f"    설명: {desc[:100] if desc else ''}")

    print("\n" + "=" * 80)
    print("[2] 해외주식 잔고 (기본 잔고조회 API)")
    print("=" * 80)

    ws2 = wb["해외주식 잔고"]

    print("\n[Request Parameters]")
    in_request_params = False
    param_rows = []

    for row in ws2.iter_rows(values_only=True):
        if row[0] and "Request Parameters" in str(row[0]):
            in_request_params = True
            continue

        if in_request_params and row[0] and "Response" in str(row[0]):
            break

        if in_request_params and row[0]:
            param_rows.append(row[:7])

    for row in param_rows[:20]:
        if row[0]:
            element = row[0] or ""
            korean = row[1] or ""
            type_val = row[2] or ""
            required = row[3] or ""
            length = row[4] or ""
            desc = row[5] or ""

            print(f"\n  - {element}")
            print(f"    한글명: {korean}")
            print(f"    타입: {type_val}")
            print(f"    필수: {required}")
            print(f"    길이: {length}")
            print(f"    설명: {desc[:100] if desc else ''}")

    print("\n" + "=" * 80)
    print("[3] 해외주식 주문 API")
    print("=" * 80)

    ws3 = wb["해외주식 주문"]

    print("\n[Request Body Parameters]")
    in_request_body = False
    param_rows = []

    for row in ws3.iter_rows(values_only=True):
        if row[0] and "Request Body" in str(row[0]):
            in_request_body = True
            continue

        if in_request_body and row[0] and "Response" in str(row[0]):
            break

        if in_request_body and row[0]:
            param_rows.append(row[:7])

    for row in param_rows[:15]:
        if row[0]:
            element = row[0] or ""
            korean = row[1] or ""
            type_val = row[2] or ""
            required = row[3] or ""
            length = row[4] or ""
            desc = row[5] or ""

            print(f"\n  - {element}")
            print(f"    한글명: {korean}")
            print(f"    타입: {type_val}")
            print(f"    필수: {required}")
            print(f"    길이: {length}")
            print(f"    설명: {desc[:100] if desc else ''}")

    wb.close()


if __name__ == "__main__":
    try:
        extract_balance_api_params()
    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()
