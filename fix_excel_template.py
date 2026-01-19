"""
Excel 템플릿 Fix #10 수정 스크립트
B11-B13에 그리드 파라미터 추가, B14-B18로 기존 설정 이동
"""
import openpyxl
from openpyxl.utils import range_boundaries

def unmerge_if_needed(ws, cell_ref):
    """셀이 병합되어 있으면 병합 해제"""
    for merged in list(ws.merged_cells.ranges):
        if cell_ref in merged:
            ws.unmerge_cells(str(merged))
            print(f"  병합 해제: {merged}")
            return True
    return False

def main():
    wb = openpyxl.load_workbook('phoenix_grid_template_v3.xlsx')
    ws = wb['01_매매전략_기준설정']

    print('=== 수정 전 셀 내용 (B11-B17) ===')
    for row in range(11, 18):
        cell = ws[f'B{row}']
        val = cell.value if not isinstance(cell, openpyxl.cell.cell.MergedCell) else '[Merged]'
        print(f'B{row}: {val}')

    # 병합 셀 확인 및 해제
    print('\n=== 병합 셀 해제 ===')
    for row in range(11, 18):
        unmerge_if_needed(ws, f'B{row}')
        unmerge_if_needed(ws, f'A{row}')

    # 기존 데이터 백업 (B13-B16 → 임시 저장)
    print('\n=== 기존 데이터 백업 ===')
    backup = {}
    for row in range(13, 17):
        cell = ws[f'B{row}']
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            backup[row] = cell.value
            print(f'  Backup B{row}: {cell.value}')

    # 새 값 설정
    print('\n=== 새 값 설정 ===')

    # A열 레이블
    ws['A11'] = '매수간격 (%)'
    ws['A12'] = '매도목표 (%)'
    ws['A13'] = '시드비율 (%)'

    # B열 값
    ws['B11'] = 0.005  # buy_interval: 0.5%
    ws['B12'] = 0.03   # sell_target: 3%
    ws['B13'] = 0.05   # seed_ratio: 5%

    print('  B11: 0.005 (매수간격 0.5%)')
    print('  B12: 0.03 (매도목표 3%)')
    print('  B13: 0.05 (시드비율 5%)')

    # 기존 데이터 한 칸씩 아래로 이동
    print('\n=== 기존 텔레그램 설정 이동 (B13-B16 → B14-B17) ===')
    if 13 in backup:
        ws['B14'] = backup[13]  # Telegram ID
        ws['A14'] = '텔레그램 ID'
        print(f'  B14: {backup[13]}')

    if 14 in backup:
        ws['B15'] = backup[14]  # Telegram Token
        ws['A15'] = '텔레그램 Token'
        print(f'  B15: {backup[14]}')

    if 15 in backup:
        ws['B16'] = backup[15]  # Telegram Enabled
        ws['A16'] = '텔레그램 알림'
        print(f'  B16: {backup[15]}')

    if 16 in backup:
        ws['B17'] = backup[16]  # Excel Update Interval
        ws['A17'] = 'Excel 업데이트 주기 (초)'
        print(f'  B17: {backup[16]}')

    # 확인
    print('\n=== 수정 후 셀 내용 (B11-B18) ===')
    for row in range(11, 19):
        cell = ws[f'B{row}']
        val = cell.value if not isinstance(cell, openpyxl.cell.cell.MergedCell) else '[Merged]'
        print(f'B{row}: {val}')

    # 저장
    wb.save('phoenix_grid_template_v3.xlsx')
    wb.close()

    print('\n✅ Excel 템플릿 수정 완료: phoenix_grid_template_v3.xlsx')

if __name__ == '__main__':
    main()
