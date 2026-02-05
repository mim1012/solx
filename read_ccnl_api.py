# -*- coding: utf-8 -*-
"""주문체결내역 API 상세 읽기"""
import openpyxl

file_path = r"C:\Users\PC_1M\Downloads\[해외주식] 주문_계좌.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

# "해외주식 주문체결내역" 시트 읽기
sheet_name = None
for name in wb.sheetnames:
    if '주문체결' in name or 'ü' in name:
        sheet_name = name
        break

if sheet_name:
    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    print("=" * 100)

    for row_idx in range(1, min(150, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(15, ws.max_column + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                row_data.append(str(cell_value))

        if row_data:
            row_text = " | ".join(row_data)
            print(f"Row {row_idx}: {row_text}")
else:
    print("Sheet not found!")

wb.close()
