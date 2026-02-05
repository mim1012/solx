# -*- coding: utf-8 -*-
"""KIS API 문서 엑셀 파일 읽기"""
import openpyxl
import json

file_path = r"C:\Users\PC_1M\Downloads\[해외주식] 주문_계좌.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

print("=" * 80)
print("Sheet Names:")
print("=" * 80)
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

print("\n" + "=" * 80)
print("Analyzing sheets for order/execution related APIs...")
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # 각 시트의 첫 50행 정도를 스캔하여 API 이름과 설명 찾기
    print(f"\n[Sheet: {sheet_name}]")
    print("-" * 80)

    for row_idx in range(1, min(100, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                row_data.append(str(cell_value))

        if row_data:
            row_text = " | ".join(row_data)
            # 체결, 주문, 조회 관련 키워드 찾기
            keywords = ['체결', '주문', '조회', 'output', 'Output', 'OUTPUT', 'AVG_PRVS', 'TOT_CCLD', 'ODNO']
            if any(keyword in row_text for keyword in keywords):
                print(f"Row {row_idx}: {row_text[:200]}")

wb.close()
