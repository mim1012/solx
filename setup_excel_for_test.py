#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 템플릿 초기 설정 스크립트
소액 테스트용 설정 값 입력
"""

from openpyxl import load_workbook

def setup_excel_for_test():
    """소액 테스트용 Excel 설정"""

    wb = load_workbook('phoenix_grid_template_v3.xlsx')
    ws = wb.active

    # 소액 테스트 설정
    print("[1/8] Setting B15 (System Active) = FALSE")
    ws['B15'] = False  # 처음엔 FALSE (설정 검증)

    print("[2/8] Setting B16 (Ticker) = SOXL")
    ws['B16'] = "SOXL"

    print("[3/8] Setting B17 (Initial Capital) = 500")
    ws['B17'] = 500  # 소액 테스트: $500

    print("[4/8] Setting B18 (Tier 1) = None (auto-detect)")
    ws['B18'] = None  # 자동 설정

    print("[5/8] Setting B19 (Tier Amount) = 50")
    ws['B19'] = 50  # 티어당 $50

    print("[6/8] Setting B20 (Tier 1 Custom) = FALSE")
    ws['B20'] = False  # 갭 상승 모드 OFF

    print("[7/8] Setting B21 (Price Check Interval) = 40")
    ws['B21'] = 40  # 40초 주기

    print("[8/8] Setting B22 (Fill Check Enabled) = TRUE")
    ws['B22'] = True  # 체결 확인 ON (필수)

    # 저장
    wb.save('phoenix_grid_template_v3.xlsx')
    print("\n=== Excel Template Setup Complete ===")
    print("WARNING: B12 (APP KEY), B13 (APP SECRET), B14 (Account) need manual input!")
    print("Please open Excel and enter your KIS API credentials.")
    print("\nCurrent settings:")
    print(f"  B15 (System Active): {ws['B15'].value}")
    print(f"  B16 (Ticker): {ws['B16'].value}")
    print(f"  B17 (Capital): ${ws['B17'].value}")
    print(f"  B19 (Tier Amount): ${ws['B19'].value}")
    print(f"  B21 (Interval): {ws['B21'].value}s")
    print(f"  B22 (Fill Check): {ws['B22'].value}")

if __name__ == "__main__":
    setup_excel_for_test()
