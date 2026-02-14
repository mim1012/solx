"""
Phoenix Excel Bridge - Excel 파일 읽기/쓰기 인터페이스 v3.1
2-sheet 구조 지원 (01_매매전략_기준설정, 02_운용로그_히스토리)
"""
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional
from datetime import datetime
import logging
import time

from .models import GridSettings, Position, SystemState


logger = logging.getLogger(__name__)


class ExcelBridge:
    """
    Excel 파일과 그리드 엔진 간 데이터 브리지

    시트 1: "01_매매전략_기준설정"
    - 영역 A (A1:B16): 기본 설정 읽기
    - 영역 B (D1:E10): 프로그램 매매 정보 쓰기
    - 영역 C (A17:E257): 240개 티어 테이블 (읽기 전용)
    - 영역 D (G17:N257): 프로그램 시뮬레이션 영역 (쓰기)

    시트 2: "02_운용로그_히스토리"
    - 시간 순 누적 로그 (append-only)
    """

    def __init__(self, file_path: str):
        """
        Excel 브리지 초기화

        Args:
            file_path: Excel 파일 경로
        """
        self.file_path = file_path
        self.wb = None
        self.ws_master = None
        self.ws_history = None
        self.ws_config = None  # [v4.2] 시스템설정 시트

    def load_workbook(self):
        """Excel 파일 열기"""
        try:
            self.wb = openpyxl.load_workbook(self.file_path)
            self.ws_master = self.wb["01_매매전략_기준설정"]
            self.ws_history = self.wb["02_운용로그_히스토리"]

            # [v4.2] 시스템설정 시트 (선택사항, 없으면 None)
            if "03_시스템설정" in self.wb.sheetnames:
                self.ws_config = self.wb["03_시스템설정"]
                logger.info(f"Excel 파일 로드 성공 (시스템설정 시트 포함): {self.file_path}")
            else:
                self.ws_config = None
                logger.info(f"Excel 파일 로드 성공 (시스템설정 시트 없음): {self.file_path}")
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel 파일을 찾을 수 없음: {self.file_path}")
        except KeyError as e:
            raise KeyError(f"필수 시트를 찾을 수 없음: {e}")

    def save_workbook(self, max_retries: int = 3, retry_delay: float = 1.0) -> bool:
        """
        Excel 파일 저장 (재시도 로직 포함)

        Args:
            max_retries: 최대 재시도 횟수
            retry_delay: 재시도 대기 시간 (초)

        Returns:
            bool: 저장 성공 여부
        """
        if not self.wb:
            logger.warning("Workbook이 로드되지 않음")
            return False

        for attempt in range(max_retries):
            try:
                self.wb.save(self.file_path)
                if attempt > 0:
                    logger.info(f"Excel 파일 저장 성공 (재시도 {attempt}회 후): {self.file_path}")
                else:
                    logger.info(f"Excel 파일 저장 완료: {self.file_path}")
                return True
            except PermissionError as e:
                if attempt < max_retries - 1:
                    logger.warning(f"Excel 파일 잠금 감지, {retry_delay}초 후 재시도 ({attempt + 1}/{max_retries}): {e}")
                    time.sleep(retry_delay)
                else:
                    logger.error(f"Excel 파일 저장 실패 (최대 재시도 초과): {e}")
                    return False
            except Exception as e:
                logger.error(f"Excel 파일 저장 중 예기치 않은 오류: {e}")
                return False

        return False

    def close_workbook(self):
        """Excel 파일 닫기"""
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws_master = None
            self.ws_history = None
            self.ws_config = None  # [v4.2] 시스템설정 시트

    def _read_bool(self, cell) -> bool:
        """
        Excel 셀에서 boolean 값을 안전하게 읽음

        True: 1, "1", "true", "yes", "y", "on"
        False: 0, "0", "false", "no", "n", "off", None, ""

        Args:
            cell: openpyxl cell 객체

        Returns:
            boolean 값
        """
        if cell.value is None:
            return False

        if isinstance(cell.value, bool):
            return cell.value

        if isinstance(cell.value, (int, float)):
            return cell.value != 0

        if isinstance(cell.value, str):
            raw = cell.value.strip().lower()
            true_values = {"1", "true", "y", "yes", "t", "on"}
            false_values = {"0", "false", "n", "no", "f", "off", ""}

            if raw in true_values:
                return True
            elif raw in false_values:
                return False
            else:
                logger.warning(f"Unknown boolean value: '{cell.value}', defaulting to False")
                return False

        return False

    def _parse_percent(self, cell, default: float) -> float:
        """
        퍼센트 값 안전 파싱 (10% → 0.10)

        Args:
            cell: openpyxl cell 객체
            default: 기본값

        Returns:
            파싱된 퍼센트 값 (소수)
        """
        try:
            value = cell.value
            if value is None or value == "":
                return default

            # 문자열 퍼센트 처리
            if isinstance(value, str):
                value = value.strip().replace("%", "").replace(",", "")

            result = float(value)

            # 1 이상이면 퍼센트로 가정 (10% → 0.10)
            if result >= 1:
                result = result / 100

            return result
        except Exception as e:
            logger.warning(f"퍼센트 파싱 실패 {cell.coordinate}: {e}, 기본값 {default} 사용")
            return default

    def load_settings(self) -> GridSettings:
        """
        시트 1 "01_매매전략_기준설정" 영역 A에서 매매 파라미터 읽기

        Returns:
            그리드 설정 객체
        """
        if not self.ws_master:
            self.load_workbook()

        # 기본 설정 (A1:B10)
        account_no = self.ws_master["B2"].value
        ticker = self.ws_master["B3"].value
        investment_usd = float(self.ws_master["B4"].value or 0)
        total_tiers = int(self.ws_master["B5"].value or 240)

        # [FIX] B6 (티어당 투자금) 자동 계산
        tier_amount_raw = self.ws_master["B6"].value

        # B6이 수식(=로 시작), None, 0이면 자동 계산
        if (tier_amount_raw is None or
            tier_amount_raw == 0 or
            (isinstance(tier_amount_raw, str) and tier_amount_raw.startswith("="))):
            # 자동 계산 (B4/B5)
            tier_amount = investment_usd / total_tiers if total_tiers > 0 else 0
            logger.info(f"[AUTO] 티어당 투자금 자동 계산: ${tier_amount:.2f} = ${investment_usd:,.0f} / {total_tiers} tiers")
        else:
            tier_amount = float(tier_amount_raw)
            logger.info(f"[MANUAL] 티어당 투자금 Excel 설정값 사용: ${tier_amount:.2f}")

        tier1_auto_update = self._read_bool(self.ws_master["B7"])

        # [CUSTOM v3.1] Tier 1 거래 설정
        tier1_trading_enabled = self._read_bool(self.ws_master["B8"])

        buy_limit = self._read_bool(self.ws_master["B9"])
        sell_limit = self._read_bool(self.ws_master["B10"])

        # [v4.1] KIS API 설정 (B12~B15)
        kis_app_key = self.ws_master["B12"].value or ""
        kis_app_secret = self.ws_master["B13"].value or ""
        kis_account_no = self.ws_master["B14"].value or ""
        system_running = self._read_bool(self.ws_master["B15"])

        # 그리드 파라미터 (기본값 유지, 향후 추가 가능)
        buy_interval = 0.005  # 0.5%
        sell_target = 0.03    # 3%
        seed_ratio = 0.05     # 5%

        # 텔레그램 알림 (행 번호 변경: B18~B20)
        telegram_id = self.ws_master["B18"].value
        telegram_token = self.ws_master["B19"].value
        telegram_enabled = self._read_bool(self.ws_master["B20"])

        # 시스템 설정
        price_check_interval = float(self.ws_master["B16"].value or 40.0)  # 시세 조회 주기 (초)
        excel_update_interval = float(self.ws_master["B21"].value or 1.0)

        # [v4.2] 장시간 설정 (03_시스템설정 시트에서 읽기)
        # None이면 config.py의 MARKET_HOURS_EDT/EST를 사용
        edt_regular_open_hour = None
        edt_regular_open_minute = None
        edt_regular_close_hour = None
        edt_regular_close_minute = None
        edt_premarket_start = None
        edt_aftermarket_end = None
        est_regular_open_hour = None
        est_regular_open_minute = None
        est_regular_close_hour = None
        est_regular_close_minute = None
        est_premarket_start = None
        est_aftermarket_end = None
        enable_premarket = None
        enable_aftermarket = None

        # 03_시스템설정 시트가 있으면 설정 읽기
        # 셀 매핑: B8-B13 (EDT), B16-B21 (EST), B24-B25 (옵션)
        if self.ws_config is not None:
            edt_regular_open_hour = self.ws_config["B8"].value
            edt_regular_open_minute = self.ws_config["B9"].value
            edt_regular_close_hour = self.ws_config["B10"].value
            edt_regular_close_minute = self.ws_config["B11"].value
            edt_premarket_start = self.ws_config["B12"].value
            edt_aftermarket_end = self.ws_config["B13"].value
            est_regular_open_hour = self.ws_config["B16"].value
            est_regular_open_minute = self.ws_config["B17"].value
            est_regular_close_hour = self.ws_config["B18"].value
            est_regular_close_minute = self.ws_config["B19"].value
            est_premarket_start = self.ws_config["B20"].value
            est_aftermarket_end = self.ws_config["B21"].value
            enable_premarket = self._read_bool(self.ws_config["B24"]) if self.ws_config["B24"].value is not None else None
            enable_aftermarket = self._read_bool(self.ws_config["B25"]) if self.ws_config["B25"].value is not None else None

            # 정수형 변환 (None이 아닌 경우에만)
            edt_regular_open_hour = int(edt_regular_open_hour) if edt_regular_open_hour is not None else None
            edt_regular_open_minute = int(edt_regular_open_minute) if edt_regular_open_minute is not None else None
            edt_regular_close_hour = int(edt_regular_close_hour) if edt_regular_close_hour is not None else None
            edt_regular_close_minute = int(edt_regular_close_minute) if edt_regular_close_minute is not None else None
            edt_premarket_start = int(edt_premarket_start) if edt_premarket_start is not None else None
            edt_aftermarket_end = int(edt_aftermarket_end) if edt_aftermarket_end is not None else None
            est_regular_open_hour = int(est_regular_open_hour) if est_regular_open_hour is not None else None
            est_regular_open_minute = int(est_regular_open_minute) if est_regular_open_minute is not None else None
            est_regular_close_hour = int(est_regular_close_hour) if est_regular_close_hour is not None else None
            est_regular_close_minute = int(est_regular_close_minute) if est_regular_close_minute is not None else None
            est_premarket_start = int(est_premarket_start) if est_premarket_start is not None else None
            est_aftermarket_end = int(est_aftermarket_end) if est_aftermarket_end is not None else None

            logger.info(f"[v4.2] 시스템설정 시트에서 장시간 설정 로드 완료")
        else:
            logger.info(f"[v4.2] 시스템설정 시트 없음, config.py 기본값 사용")

        # [CUSTOM v3.1] Tier 1 매수% 읽기 (영역 C, 행 18, 컬럼 C)
        tier1_buy_percent = float(self.ws_master["C18"].value or 0.0)

        # 검증
        if ticker != "SOXL":
            raise ValueError(f"지원하지 않는 종목: {ticker}. SOXL만 지원합니다.")

        # [CUSTOM v3.1] Tier 1 거래 모드 검증
        if tier1_trading_enabled and tier1_buy_percent is None:
            raise ValueError("Tier 1 거래 활성화 시 tier1_buy_percent 설정 필수")

        settings = GridSettings(
            account_no=account_no or "",
            ticker=ticker,
            investment_usd=investment_usd,
            total_tiers=total_tiers,
            tier_amount=tier_amount,
            tier1_auto_update=tier1_auto_update,
            tier1_trading_enabled=tier1_trading_enabled,
            tier1_buy_percent=tier1_buy_percent,
            buy_limit=buy_limit,
            sell_limit=sell_limit,
            # [v4.1] KIS API 설정
            kis_app_key=kis_app_key,
            kis_app_secret=kis_app_secret,
            kis_account_no=kis_account_no,
            system_running=system_running,
            # 텔레그램 알림
            telegram_id=telegram_id,
            telegram_token=telegram_token,
            telegram_enabled=telegram_enabled,
            excel_update_interval=excel_update_interval,
            price_check_interval=price_check_interval,
            # 그리드 파라미터 추가
            seed_ratio=seed_ratio,
            buy_interval=buy_interval,
            sell_target=sell_target,
            # [v4.2] 장시간 설정 (Excel에서 읽음, None이면 config.py 사용)
            edt_regular_open_hour=edt_regular_open_hour,
            edt_regular_open_minute=edt_regular_open_minute,
            edt_regular_close_hour=edt_regular_close_hour,
            edt_regular_close_minute=edt_regular_close_minute,
            edt_premarket_start=edt_premarket_start,
            edt_aftermarket_end=edt_aftermarket_end,
            est_regular_open_hour=est_regular_open_hour,
            est_regular_open_minute=est_regular_open_minute,
            est_regular_close_hour=est_regular_close_hour,
            est_regular_close_minute=est_regular_close_minute,
            est_premarket_start=est_premarket_start,
            est_aftermarket_end=est_aftermarket_end,
            enable_premarket=enable_premarket,
            enable_aftermarket=enable_aftermarket
        )

        logger.info(f"설정 로드 완료: {ticker}, 투자금=${investment_usd:.2f}, "
                   f"Tier 1 거래={'ON' if tier1_trading_enabled else 'OFF'}, "
                   f"buy_interval={buy_interval:.3%}, sell_target={sell_target:.2%}")

        return settings

    def load_tier_table(self) -> List[Dict]:
        """
        시트 1 영역 C (A23:E262) 240개 티어 테이블 읽기
        [P0 FIX] Row 22는 헤더, 실제 데이터는 23~262행

        Returns:
            티어 테이블 데이터 (리스트)
        """
        if not self.ws_master:
            self.load_workbook()

        tier_table = []
        for row_idx in range(23, 263):  # 23~262행 (240개)
            tier = self.ws_master.cell(row=row_idx, column=1).value
            seed_pct = self.ws_master.cell(row=row_idx, column=2).value
            buy_pct = self.ws_master.cell(row=row_idx, column=3).value
            sell_pct = self.ws_master.cell(row=row_idx, column=4).value
            remark = self.ws_master.cell(row=row_idx, column=5).value

            tier_table.append({
                "tier": int(tier) if tier else 0,
                "seed_pct": float(seed_pct) if seed_pct else 0.0,
                "buy_pct": float(buy_pct) if buy_pct else 0.0,
                "sell_pct": float(sell_pct) if sell_pct else 0.0,
                "remark": remark or ""
            })

        logger.info(f"티어 테이블 로드 완료: {len(tier_table)}개 티어")
        return tier_table

    def update_program_info(self, state: SystemState):
        """
        시트 1 영역 B (D1:E10) 프로그램 매매 정보 업데이트

        Args:
            state: 시스템 현재 상태
        """
        if not self.ws_master:
            self.load_workbook()

        # 최근 업데이트 시간
        self.ws_master["E2"].value = state.last_update.strftime("%Y-%m-%d %H:%M:%S")

        # 현재 티어
        self.ws_master["E3"].value = state.current_tier

        # 현재가
        self.ws_master["E4"].value = state.current_price

        # 잔고
        self.ws_master["E5"].value = state.account_balance

        # 수량차 (총 보유 수량)
        self.ws_master["E6"].value = state.total_quantity

        # 매수/매도 상태
        self.ws_master["E7"].value = state.buy_status
        self.ws_master["E8"].value = state.sell_status

        logger.debug(f"프로그램 정보 업데이트: Tier {state.current_tier}, ${state.current_price:.2f}")

    def update_program_area(self, positions: List[Position], tier1_price: float, buy_interval: float = 0.005):
        """
        시트 1 영역 D (G17:N257) 프로그램 시뮬레이션 영역 업데이트

        컬럼: G=티어, H=잔고량, I=투자금, J=티어평단, K=매수(가), L=매수(량), M=매도(가), N=매도(량)

        Args:
            positions: 현재 보유 포지션 리스트
            tier1_price: Tier 1 기준가
            buy_interval: 티어 간 하락 간격 (기본값 0.005 = 0.5%)
        """
        if not self.ws_master:
            self.load_workbook()

        # 240개 티어 초기화
        for tier in range(1, 241):
            row_idx = 17 + tier  # 18~257행

            # 티어 번호
            self.ws_master.cell(row=row_idx, column=7, value=tier)

            # 해당 티어 포지션 찾기
            position = next((p for p in positions if p.tier == tier), None)

            if position:
                # 보유 중인 티어
                self.ws_master.cell(row=row_idx, column=8, value=position.quantity)  # 잔고량
                self.ws_master.cell(row=row_idx, column=9, value=position.invested_amount)  # 투자금
                self.ws_master.cell(row=row_idx, column=10, value=position.avg_price)  # 티어평단

                # 매수(가), 매수(량) - 이미 보유 중이므로 빈값
                self.ws_master.cell(row=row_idx, column=11, value="")
                self.ws_master.cell(row=row_idx, column=12, value="")

                # 매도(가), 매도(량) - 목표가 계산
                sell_price = position.avg_price * 1.03  # 3% 익절
                self.ws_master.cell(row=row_idx, column=13, value=sell_price)
                self.ws_master.cell(row=row_idx, column=14, value=position.quantity)
            else:
                # 미보유 티어 - 매수 대기 상태
                self.ws_master.cell(row=row_idx, column=8, value=0)  # 잔고량
                self.ws_master.cell(row=row_idx, column=9, value=0)  # 투자금
                self.ws_master.cell(row=row_idx, column=10, value=0)  # 티어평단

                # 매수(가), 매수(량) - 예상 매수가 계산
                if tier == 1:
                    buy_price = tier1_price
                else:
                    decline_rate = (tier - 1) * buy_interval
                    buy_price = tier1_price * (1 - decline_rate)

                # 예상 매수 수량 (tier_amount는 settings에서 가져와야 하므로 생략, 실제 구현 시 전달 필요)
                self.ws_master.cell(row=row_idx, column=11, value=buy_price)
                self.ws_master.cell(row=row_idx, column=12, value="")  # 수량은 체결 시 업데이트

                # 매도(가), 매도(량) - 빈값
                self.ws_master.cell(row=row_idx, column=13, value="")
                self.ws_master.cell(row=row_idx, column=14, value="")

        logger.debug(f"프로그램 영역 업데이트 완료: {len(positions)}개 포지션")

    def append_history_log(self, log_entry: Dict):
        """
        시트 2 "02_운용로그_히스토리"에 로그 추가 (append-only)

        Args:
            log_entry: 로그 엔트리 (dict)
                - update_time: 업데이트 시간
                - date: 날짜
                - sheet: 시트명 ("Main")
                - ticker: 종목코드
                - tier: 현재 티어
                - total_tiers: 총 티어
                - quantity_diff: 수량차
                - invested: 투자금
                - tier_amount: 1티어 금액
                - balance: 예수금
                - stock_value: 주식평가금
                - holding_profit: 잔고수익
                - buy_ready: 매수예정
                - withdrawable: 인출가능
                - arbitrage_profit: 아비타수익
                - buy_qty: 매수 수량
                - sell_qty: 매도 수량
        """
        if not self.ws_history:
            self.load_workbook()

        # 다음 빈 행 찾기
        next_row = self.ws_history.max_row + 1

        # A. 시간·기본 정보
        self.ws_history.cell(row=next_row, column=1, value=log_entry.get("update_time"))  # 업데이트
        self.ws_history.cell(row=next_row, column=2, value=log_entry.get("date"))  # 날짜
        self.ws_history.cell(row=next_row, column=3, value=log_entry.get("sheet", "Main"))  # 시트
        self.ws_history.cell(row=next_row, column=4, value=log_entry.get("ticker", "SOXL"))  # 종목
        self.ws_history.cell(row=next_row, column=5, value=log_entry.get("tier"))  # 티어

        # B. 포지션 상태
        self.ws_history.cell(row=next_row, column=6, value=log_entry.get("total_tiers"))  # 총티어
        self.ws_history.cell(row=next_row, column=7, value=log_entry.get("quantity_diff"))  # 잔고량(차)
        self.ws_history.cell(row=next_row, column=8, value=log_entry.get("invested"))  # 투자금
        self.ws_history.cell(row=next_row, column=9, value=log_entry.get("tier_amount"))  # 1티어

        # C. 손익·성과 지표
        self.ws_history.cell(row=next_row, column=10, value=log_entry.get("balance"))  # 예수금
        self.ws_history.cell(row=next_row, column=11, value=log_entry.get("stock_value"))  # 주식평가금
        self.ws_history.cell(row=next_row, column=12, value=log_entry.get("holding_profit"))  # 잔고수익
        self.ws_history.cell(row=next_row, column=13, value=log_entry.get("buy_ready"))  # 매수예정
        self.ws_history.cell(row=next_row, column=14, value=log_entry.get("withdrawable"))  # 인출가능
        self.ws_history.cell(row=next_row, column=15, value=log_entry.get("arbitrage_profit"))  # 아비타수익

        # D. 실제 체결 수량
        self.ws_history.cell(row=next_row, column=16, value=log_entry.get("buy_qty", 0))  # 매수
        self.ws_history.cell(row=next_row, column=17, value=log_entry.get("sell_qty", 0))  # 매도

        logger.info(f"히스토리 로그 추가: 행 {next_row}, Tier {log_entry.get('tier')}")

    def create_history_log_entry(self, state: SystemState,
                                 settings: GridSettings,
                                 buy_qty: int = 0,
                                 sell_qty: int = 0) -> Dict:
        """
        히스토리 로그 엔트리 생성 헬퍼 함수

        Args:
            state: 시스템 상태
            settings: 그리드 설정
            buy_qty: 매수 수량 (최근 체결)
            sell_qty: 매도 수량 (최근 체결)

        Returns:
            로그 엔트리 (dict)
        """
        now = datetime.now()

        # 아비타 수익 계산 (간단 버전: 수익률 기준)
        arbitrage_profit = state.profit_rate

        # 매수 예정 금액 (1티어 금액)
        buy_ready = settings.tier_amount if state.account_balance >= settings.tier_amount else 0

        # 인출 가능 금액 (예수금 - 1티어 금액)
        withdrawable = max(0, state.account_balance - settings.tier_amount)

        log_entry = {
            "update_time": now.strftime("%Y-%m-%d %H:%M:%S"),
            "date": now.strftime("%Y-%m-%d"),
            "sheet": "Main",
            "ticker": settings.ticker,
            "tier": state.current_tier,
            "total_tiers": settings.total_tiers,
            "quantity_diff": state.total_quantity,
            "invested": state.total_invested,
            "tier_amount": settings.tier_amount,
            "balance": state.account_balance,
            "stock_value": state.stock_value,
            "holding_profit": state.profit_rate,  # 수익률로 저장
            "buy_ready": buy_ready,
            "withdrawable": withdrawable,
            "arbitrage_profit": arbitrage_profit,
            "buy_qty": buy_qty,
            "sell_qty": sell_qty
        }

        return log_entry
