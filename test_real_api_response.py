"""실제 USD 예수금 API 응답 확인 스크립트"""
import sys
import logging
from openpyxl import load_workbook
from src.kis_rest_adapter import KisRestAdapter

# 로깅 설정
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)

logger = logging.getLogger(__name__)

print("=" * 80)
print("USD 예수금 API 응답 확인 (실제 데이터)")
print("=" * 80)

# 1. Excel에서 API 키 로드
excel_path = "release/phoenix_grid_template_v3.xlsx"
wb = load_workbook(excel_path, data_only=False)
ws = wb.active

kis_app_key = ws["B12"].value
kis_app_secret = ws["B13"].value
kis_account_no = ws["B14"].value

print(f"\n[1] API 키 로드 완료")
print(f"  - APP_KEY: {str(kis_app_key)[:4]}****")
print(f"  - ACCOUNT: {kis_account_no}")

# 2. KIS API 어댑터 초기화
adapter = KisRestAdapter(
    app_key=kis_app_key,
    app_secret=kis_app_secret,
    account_no=kis_account_no
)

# 3. 로그인
print(f"\n[2] KIS API 로그인...")
try:
    adapter.login()
    print("  [OK] 로그인 성공")
except Exception as e:
    print(f"  [FAIL] 로그인 실패: {e}")
    sys.exit(1)

# 4. SOXL 시세 조회
ticker = "SOXL"
print(f"\n[3] {ticker} 시세 조회...")
try:
    price_data = adapter.get_overseas_price(ticker)
    if not price_data:
        print(f"  [FAIL] 시세 조회 실패")
        sys.exit(1)

    current_price = price_data['price']
    print(f"  [OK] 현재가: ${current_price:.2f}")
except Exception as e:
    print(f"  [FAIL] 시세 조회 실패: {e}")
    sys.exit(1)

# 5. USD 예수금 조회 (get_cash_balance - TTTS3007R API)
print(f"\n[4] USD 예수금 조회 (TTTS3007R API)")
print(f"  - 요청 파라미터:")
print(f"    * ticker: {ticker}")
print(f"    * price: ${current_price:.2f}")
print()

try:
    cash_balance = adapter.get_cash_balance(ticker=ticker, price=current_price)

    print(f"\n[결과] USD 예수금: ${cash_balance:,.2f}")

    if cash_balance == 0.0:
        print("\n[WARNING] USD 예수금이 $0.00입니다.")
        print("위의 [DEBUG] 로그에서 API 응답을 확인하세요:")
        print("  - rt_cd: API 응답 코드 (0=성공, 7=상품없음)")
        print("  - ord_psbl_frcr_amt: 주문가능외화금액 필드")
    else:
        print(f"\n[OK] USD 예수금: ${cash_balance:,.2f}")

except Exception as e:
    print(f"\n[FAIL] 예수금 조회 실패: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n" + "=" * 80)
print("API 응답 확인 완료")
print("=" * 80)
