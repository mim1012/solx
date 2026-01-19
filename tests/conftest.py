"""
Phoenix 자동매매 시스템 pytest 설정 및 공통 fixture

코드 리뷰에서 식별된 주요 이슈들을 테스트:
1. Excel 파일 동시 접근 및 lock 처리
2. 주문 실패 시 GridEngine 상태 일관성
3. 네트워크 단절 복구
4. 실시간 시세 콜백 동시성
5. Kiwoom API 로그인 타임아웃
"""

import pytest
import tempfile
import shutil
import os
from pathlib import Path
from unittest.mock import Mock, MagicMock
import openpyxl
from datetime import datetime

# 테스트용 Excel 템플릿 경로
TEST_EXCEL_TEMPLATE = "phoenix_grid_template_v3.xlsx"


@pytest.fixture(scope="session")
def excel_template_path():
    """원본 Excel 템플릿 경로"""
    return TEST_EXCEL_TEMPLATE


@pytest.fixture
def temp_excel_file(excel_template_path):
    """테스트용 임시 Excel 파일 생성"""
    if not os.path.exists(excel_template_path):
        # 템플릿이 없으면 기본 구조 생성
        wb = openpyxl.Workbook()

        # 시트 1: 매매전략 기준설정
        ws1 = wb.active
        ws1.title = "01_매매전략_기준설정"

        # 기본 설정
        ws1["B2"] = "12345678"  # 계좌번호
        ws1["B3"] = "SOXL"      # 종목코드
        ws1["B4"] = 10000       # 투자금 USD
        ws1["B5"] = 240         # 총 티어 수
        ws1["B6"] = 50.0        # 티어당 금액
        ws1["B7"] = 0           # tier1_auto_update
        ws1["B8"] = 0           # tier1_trading_enabled
        ws1["B9"] = 0           # buy_limit
        ws1["B10"] = 0          # sell_limit
        ws1["B11"] = "0.5%"     # buy_interval
        ws1["B12"] = "3%"       # sell_target
        ws1["B13"] = "5%"       # seed_ratio
        ws1["B14"] = ""         # telegram_id
        ws1["B15"] = ""         # telegram_token
        ws1["B16"] = 0          # telegram_enabled
        ws1["B17"] = 1.0        # excel_update_interval (FIX: 숫자로 설정)
        ws1["C18"] = 0.0        # tier1_buy_percent

        ws1["E4"] = 20.0        # 현재가
        ws1["E5"] = 10.0        # Tier 1 가격
        ws1["E6"] = 0           # 수량차 (보유량)

        # Tier 헤더 (20행으로 이동)
        ws1["A20"] = "Tier"
        ws1["B20"] = "기준가"
        ws1["H20"] = "잔고량"
        ws1["I20"] = "투자금"
        ws1["J20"] = "평균단가"

        # Tier 데이터 (21행부터 240개)
        for tier in range(1, 241):
            row_idx = 20 + tier
            ws1[f"A{row_idx}"] = tier
            ws1[f"B{row_idx}"] = 10.0 - (tier * 0.05)  # 기준가
            ws1[f"H{row_idx}"] = 0  # 잔고량
            ws1[f"I{row_idx}"] = 0  # 투자금
            ws1[f"J{row_idx}"] = 0  # 평균단가

        # 시트 2: 운용로그
        ws2 = wb.create_sheet("02_운용로그_히스토리")
        ws2.append([
            "업데이트", "날짜", "시트", "종목", "티어", "총티어", "잔고량(차)",
            "투자금", "1티어", "예수금", "주식평가금", "잔고수익",
            "매수예정", "인출가능", "아비타수익", "매수", "매도"
        ])

        temp_path = tempfile.mktemp(suffix=".xlsx")
        wb.save(temp_path)
    else:
        # 템플릿 복사
        temp_path = tempfile.mktemp(suffix=".xlsx")
        shutil.copy2(excel_template_path, temp_path)

        # [FIX] B17 셀을 숫자로 강제 설정 (excel_update_interval)
        wb = openpyxl.load_workbook(temp_path)
        ws1 = wb["01_매매전략_기준설정"]
        ws1["B17"] = 1.0
        wb.save(temp_path)
        wb.close()

    yield temp_path

    # 정리
    try:
        if os.path.exists(temp_path):
            os.remove(temp_path)
    except PermissionError:
        pass  # Excel 파일이 열려있으면 무시


@pytest.fixture
def mock_kiwoom_ocx():
    """Kiwoom OpenAPI QAxWidget Mock"""
    mock = MagicMock()

    # 기본 메서드 설정
    mock.dynamicCall = Mock()
    mock.setControl = Mock(return_value=True)

    # 계좌 정보
    mock.dynamicCall.side_effect = lambda method, *args: {
        "GetLoginInfo(QString)": ["12345678", "98765432"],  # 계좌번호 목록
        "GetConnectState()": 1,  # 연결 상태
    }.get(method, None)

    return mock


@pytest.fixture
def mock_telegram_bot():
    """Telegram Bot Mock"""
    mock = MagicMock()
    mock.send_message = Mock(return_value=True)
    return mock


@pytest.fixture
def grid_settings_basic():
    """기본 Grid 설정"""
    from src.models import GridSettings

    return GridSettings(
        account_no="12345678",
        ticker="SOXL",
        investment_usd=10000.0,
        total_tiers=240,
        tier_amount=50.0,  # $10000 / 240 tiers = ~$41.67, 반올림 $50
        tier1_auto_update=False,
        tier1_trading_enabled=False,
        tier1_buy_percent=0.0,
        buy_limit=False,
        sell_limit=False,
        telegram_enabled=False,
        tier1_price=10.0,
        tier_interval=0.05
    )


@pytest.fixture
def grid_settings_tier1_enabled():
    """Tier 1 활성화된 Grid 설정"""
    from src.models import GridSettings

    return GridSettings(
        account_no="12345678",
        ticker="SOXL",
        investment_usd=10000.0,
        total_tiers=240,
        tier_amount=50.0,
        tier1_auto_update=True,
        tier1_trading_enabled=True,
        tier1_buy_percent=0.0,  # 정확히 0% (Tier 1 가격에서 매수)
        buy_limit=False,
        sell_limit=False,
        telegram_enabled=False,
        tier1_price=10.0,
        tier_interval=0.05
    )


@pytest.fixture
def sample_positions():
    """샘플 포지션 리스트"""
    from src.models import Position

    return [
        Position(
            tier=1,
            quantity=10,
            avg_price=10.0,
            invested_amount=100.0,
            opened_at=datetime.now()
        ),
        Position(
            tier=2,
            quantity=20,
            avg_price=9.95,
            invested_amount=199.0,
            opened_at=datetime.now()
        ),
        Position(
            tier=3,
            quantity=30,
            avg_price=9.90,
            invested_amount=297.0,
            opened_at=datetime.now()
        )
    ]


@pytest.fixture
def mock_qt_event_loop():
    """PyQt5 QEventLoop Mock"""
    mock = MagicMock()
    mock.exec_ = Mock()
    mock.quit = Mock()
    return mock


@pytest.fixture(autouse=True)
def clean_logs():
    """각 테스트 후 로그 정리 (선택적)"""
    yield
    # 테스트 후 정리 로직 (필요시)


# 코드 리뷰에서 식별된 critical issue 테스트를 위한 fixture들

@pytest.fixture
def locked_excel_file(temp_excel_file):
    """Excel 파일 lock 상황 시뮬레이션"""
    # 파일을 열어서 lock 상태로 만듦
    wb = openpyxl.load_workbook(temp_excel_file)
    yield temp_excel_file, wb
    wb.close()


@pytest.fixture
def network_failure_simulator():
    """네트워크 실패 시뮬레이터"""
    class NetworkFailure:
        def __init__(self):
            self.fail_count = 0
            self.max_failures = 3

        def should_fail(self):
            """연속 3회까지 실패, 이후 성공"""
            if self.fail_count < self.max_failures:
                self.fail_count += 1
                return True
            return False

        def reset(self):
            self.fail_count = 0

    return NetworkFailure()


@pytest.fixture
def concurrent_price_updates():
    """동시 시세 업데이트 시뮬레이터"""
    import threading

    class ConcurrentUpdates:
        def __init__(self):
            self.prices = [10.0, 10.1, 9.9, 10.2, 9.8]
            self.callbacks = []

        def register_callback(self, callback):
            self.callbacks.append(callback)

        def simulate_concurrent_updates(self):
            """여러 스레드에서 동시에 시세 업데이트 발생"""
            threads = []
            for price in self.prices:
                for callback in self.callbacks:
                    t = threading.Thread(target=callback, args=(price,))
                    threads.append(t)
                    t.start()

            for t in threads:
                t.join()

    return ConcurrentUpdates()
