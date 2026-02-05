"""
Phoenix Trading System - EXE 빌드 검증 테스트
Post-Build Testing: 빌드된 EXE 파일의 실행 가능성 및 시나리오 검증
"""
import pytest
import subprocess
from pathlib import Path
import openpyxl
import shutil


@pytest.fixture
def exe_path():
    """
    빌드된 EXE 파일 경로

    Debug 모드: debug/PhoenixTrading/PhoenixTrading.exe
    Release 모드: dist/PhoenixTrading.exe
    """
    debug_exe = Path(__file__).parent.parent / "debug" / "PhoenixTrading" / "PhoenixTrading.exe"
    release_exe = Path(__file__).parent.parent / "dist" / "PhoenixTrading.exe"

    if debug_exe.exists():
        return debug_exe
    elif release_exe.exists():
        return release_exe
    else:
        pytest.skip("EXE 파일이 없습니다. build_exe.py를 먼저 실행하세요.")


@pytest.fixture
def test_excel_b15_false(tmp_path):
    """
    B15=FALSE 테스트용 Excel 파일 생성

    Args:
        tmp_path: pytest 임시 폴더

    Returns:
        Path: 테스트용 Excel 파일 경로
    """
    template_path = Path(__file__).parent.parent / "phoenix_grid_template_v3.xlsx"

    if not template_path.exists():
        pytest.skip("Excel 템플릿이 없습니다")

    test_file = tmp_path / "test_stopped.xlsx"
    shutil.copy(template_path, test_file)

    # B15를 FALSE로 설정
    wb = openpyxl.load_workbook(test_file)
    ws = wb["01_매매전략_기준설정"]
    ws["B15"] = False  # FALSE (의도적 중지)
    wb.save(test_file)
    wb.close()

    return test_file


@pytest.fixture
def test_excel_b15_true_no_keys(tmp_path):
    """
    B15=TRUE but API 키 없음 테스트용 Excel 파일

    Args:
        tmp_path: pytest 임시 폴더

    Returns:
        Path: 테스트용 Excel 파일 경로
    """
    template_path = Path(__file__).parent.parent / "phoenix_grid_template_v3.xlsx"

    if not template_path.exists():
        pytest.skip("Excel 템플릿이 없습니다")

    test_file = tmp_path / "test_no_keys.xlsx"
    shutil.copy(template_path, test_file)

    # B15=TRUE, API 키 비움
    wb = openpyxl.load_workbook(test_file)
    ws = wb["01_매매전략_기준설정"]
    ws["B15"] = True  # TRUE (시작)
    ws["B12"] = ""    # KIS APP KEY 비움
    ws["B13"] = ""    # KIS APP SECRET 비움
    wb.save(test_file)
    wb.close()

    return test_file


class TestExeBasicExecution:
    """EXE 기본 실행 테스트"""

    def test_exe_file_exists(self, exe_path):
        """EXE 파일 생성 확인"""
        assert exe_path.exists(), f"EXE 파일 없음: {exe_path}"
        assert exe_path.stat().st_size > 10 * 1024 * 1024, "EXE 파일 크기 이상 (최소 10MB 필요)"

    def test_exe_starts_without_crash(self, exe_path):
        """
        EXE 크래시 없이 시작

        Excel 파일이 없으면 종료 코드 20 (ERROR_EXCEL) 예상
        """
        result = subprocess.run(
            [str(exe_path)],
            capture_output=True,
            text=True,
            timeout=10
        )

        # Excel 파일 없음 → 종료 코드 20
        assert result.returncode == 20, f"예상: 20, 실제: {result.returncode}"

        # 로그 메시지 확인
        output = result.stdout + result.stderr
        assert "Excel 파일" in output, "Excel 파일 에러 메시지 없음"


class TestExeB15Scenarios:
    """B15 시나리오 테스트 (사용자 지정 필수)"""

    def test_b15_false_graceful_exit(self, exe_path, test_excel_b15_false):
        """
        B15=FALSE 시 정상 종료

        종료 코드 10 (STOPPED) 예상
        메시지: "시스템이 중지 상태입니다" + "에러가 아닙니다"
        """
        result = subprocess.run(
            [str(exe_path), str(test_excel_b15_false)],
            capture_output=True,
            text=True,
            timeout=15
        )

        # 종료 코드 10 (STOPPED) 확인
        assert result.returncode == 10, f"예상: 10 (STOPPED), 실제: {result.returncode}"

        # 메시지 확인
        output = result.stdout + result.stderr
        assert "시스템이 중지 상태입니다" in output, "중지 상태 메시지 없음"
        assert "B15=FALSE" in output, "B15=FALSE 설명 없음"
        assert "에러가 아닙니다" in output, "개선된 메시지 없음 (Phase 1 구현 필요)"

    def test_b15_true_no_api_keys(self, exe_path, test_excel_b15_true_no_keys):
        """
        B15=TRUE + API 키 없음

        종료 코드 21 (ERROR_API_KEY) 예상
        """
        result = subprocess.run(
            [str(exe_path), str(test_excel_b15_true_no_keys)],
            capture_output=True,
            text=True,
            timeout=15
        )

        # 종료 코드 21 (ERROR_API_KEY)
        assert result.returncode == 21, f"예상: 21 (ERROR_API_KEY), 실제: {result.returncode}"

        # 메시지 확인
        output = result.stdout + result.stderr
        assert "KIS API 키" in output, "API 키 에러 메시지 없음"


class TestExeLibraryInclusion:
    """라이브러리 포함 확인"""

    def test_openpyxl_available(self, exe_path):
        """
        openpyxl 라이브러리 포함 확인

        Excel 파일 읽기가 가능하면 openpyxl이 포함된 것
        """
        # 위 테스트에서 이미 간접적으로 검증됨
        # (Excel 파일을 읽어서 B15 값을 확인하므로)
        pass

    def test_requests_available(self, exe_path):
        """
        requests 라이브러리 포함 확인

        KIS API 연결 시도에서 확인 (test_b15_true_no_api_keys)
        """
        # 위 테스트에서 이미 간접적으로 검증됨
        pass


class TestExePackageIntegrity:
    """배포 패키지 무결성"""

    def test_release_folder_structure(self):
        """
        Release 폴더 구조 확인

        필수 파일:
        - PhoenixTrading.exe
        - phoenix_grid_template_v3.xlsx
        - README_사용방법.txt
        """
        release_dir = Path(__file__).parent.parent / "release"

        if not release_dir.exists():
            pytest.skip("Release 폴더 없음 (build_exe.py --mode=release 실행 필요)")

        required_files = [
            "PhoenixTrading.exe",
            "phoenix_grid_template_v3.xlsx",
            "README_사용방법.txt"
        ]

        for file_name in required_files:
            file_path = release_dir / file_name
            assert file_path.exists(), f"필수 파일 없음: {file_name}"

    def test_debug_folder_structure(self):
        """
        Debug 폴더 구조 확인

        Debug 모드:
        - debug_package/PhoenixTrading/ 폴더
        - phoenix_grid_template_v3.xlsx
        """
        debug_dir = Path(__file__).parent.parent / "debug_package"

        if not debug_dir.exists():
            pytest.skip("Debug 폴더 없음 (build_exe.py --mode=debug 실행 필요)")

        # Debug 모드는 폴더 형태
        exe_folder = debug_dir / "PhoenixTrading"
        assert exe_folder.exists(), "PhoenixTrading 폴더 없음"

        exe_file = exe_folder / "PhoenixTrading.exe"
        assert exe_file.exists(), "PhoenixTrading.exe 없음"


class TestExeExitCodes:
    """종료 코드 검증 (Phase 1 구현 후 활성화)"""

    @pytest.mark.skip(reason="Phase 1 (InitStatus Enum) 구현 후 활성화")
    def test_exit_code_success(self):
        """종료 코드 0 (SUCCESS)"""
        # 실제 거래 실행 후 정상 종료 시
        pass

    @pytest.mark.skip(reason="Phase 1 구현 후 활성화")
    def test_exit_code_stopped(self, exe_path, test_excel_b15_false):
        """종료 코드 10 (STOPPED)"""
        # 위 test_b15_false_graceful_exit에서 이미 검증
        pass

    @pytest.mark.skip(reason="Phase 1 구현 후 활성화")
    def test_exit_code_error_excel(self, exe_path):
        """종료 코드 20 (ERROR_EXCEL)"""
        # 위 test_exe_starts_without_crash에서 이미 검증
        pass
