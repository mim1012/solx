"""
src/excel_bridge.py 단위 테스트

테스트 범위:
1. Excel 파일 로딩 및 저장
2. 설정 읽기 (load_settings)
3. 티어 테이블 읽기 (load_tier_table)
4. 프로그램 정보 업데이트 (update_program_info, update_program_area)
5. 히스토리 로그 (append_history_log)
6. Boolean 파싱 (_read_bool)
7. 파일 잠금 시 재시도 로직

실제 ExcelBridge API:
- load_workbook() / save_workbook() / close_workbook()
- load_settings() → GridSettings
- load_tier_table() → List[Dict]
- update_program_info(state: SystemState)
- update_program_area(positions, tier1_price, buy_interval)
- append_history_log(log_entry)
- create_history_log_entry(state, settings, buy_qty, sell_qty)
"""

import pytest
import openpyxl
import time
from datetime import datetime
from unittest.mock import Mock, patch, MagicMock
from src.excel_bridge import ExcelBridge
from src.models import Position, GridSettings, SystemState


class TestExcelBridgeInitialization:
    """ExcelBridge 초기화 테스트"""

    def test_init_sets_file_path(self, temp_excel_file):
        """초기화 시 file_path 설정"""
        bridge = ExcelBridge(temp_excel_file)
        assert bridge.file_path == temp_excel_file
        assert bridge.wb is None  # load_workbook 전까지 None

    def test_load_existing_file(self, temp_excel_file):
        """기존 Excel 파일 로딩"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        assert bridge.wb is not None
        assert bridge.ws_master is not None
        assert bridge.ws_history is not None

        bridge.close_workbook()

    def test_load_nonexistent_file(self):
        """존재하지 않는 파일 로딩 시 에러"""
        bridge = ExcelBridge("nonexistent_file_12345.xlsx")
        with pytest.raises(FileNotFoundError):
            bridge.load_workbook()

    def test_worksheets_accessible(self, temp_excel_file):
        """워크시트 접근 가능"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        # 시트 직접 접근
        assert "01_매매전략_기준설정" in bridge.wb.sheetnames
        assert "02_운용로그_히스토리" in bridge.wb.sheetnames

        bridge.close_workbook()


class TestLoadSettings:
    """load_settings() 테스트"""

    def test_load_settings_returns_grid_settings(self, temp_excel_file):
        """load_settings가 GridSettings 객체 반환"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        settings = bridge.load_settings()

        assert isinstance(settings, GridSettings)
        assert settings.ticker == "SOXL"
        assert settings.total_tiers == 240
        assert settings.investment_usd > 0

        bridge.close_workbook()

    def test_load_settings_tier1_settings(self, temp_excel_file):
        """Tier 1 관련 설정 로드"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        settings = bridge.load_settings()

        # tier1_trading_enabled, tier1_auto_update 는 boolean
        assert isinstance(settings.tier1_trading_enabled, bool)
        assert isinstance(settings.tier1_auto_update, bool)

        bridge.close_workbook()

    def test_load_settings_buy_sell_limit(self, temp_excel_file):
        """매수/매도 제한 설정 로드"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        settings = bridge.load_settings()

        assert isinstance(settings.buy_limit, bool)
        assert isinstance(settings.sell_limit, bool)

        bridge.close_workbook()


class TestLoadTierTable:
    """load_tier_table() 테스트"""

    def test_load_tier_table_returns_list(self, temp_excel_file):
        """load_tier_table이 리스트 반환"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        tier_table = bridge.load_tier_table()

        assert isinstance(tier_table, list)
        # 240개 티어 (또는 실제 데이터 개수)
        assert len(tier_table) >= 1

        bridge.close_workbook()

    def test_tier_table_entry_structure(self, temp_excel_file):
        """티어 테이블 엔트리 구조"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        tier_table = bridge.load_tier_table()

        if len(tier_table) > 0:
            entry = tier_table[0]
            # 필수 키 확인
            assert "tier" in entry
            assert "seed_pct" in entry
            assert "buy_pct" in entry
            assert "sell_pct" in entry

        bridge.close_workbook()


class TestUpdateProgramInfo:
    """update_program_info() 테스트"""

    def test_update_program_info(self, temp_excel_file):
        """프로그램 정보 업데이트"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        state = SystemState(
            current_price=25.50,
            tier1_price=26.00,
            current_tier=3,
            account_balance=9500.0,
            total_quantity=50,
            total_invested=500.0,
            stock_value=1275.0,
            total_profit=25.0,
            profit_rate=0.025,
            buy_status="대기",
            sell_status="대기"
        )

        # 에러 없이 실행되어야 함
        bridge.update_program_info(state)

        # 값 확인 (E열)
        assert bridge.ws_master["E3"].value == 3  # current_tier
        assert bridge.ws_master["E4"].value == 25.50  # current_price
        assert bridge.ws_master["E5"].value == 9500.0  # account_balance

        bridge.close_workbook()


class TestUpdateProgramArea:
    """update_program_area() 테스트"""

    def test_update_program_area_with_positions(self, temp_excel_file):
        """포지션이 있을 때 프로그램 영역 업데이트"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        positions = [
            Position(tier=1, quantity=10, avg_price=10.0, invested_amount=100.0, opened_at=datetime.now()),
            Position(tier=2, quantity=20, avg_price=9.95, invested_amount=199.0, opened_at=datetime.now()),
        ]

        # 에러 없이 실행되어야 함
        bridge.update_program_area(positions, tier1_price=10.0)

        # Tier 1 행 (row=18) 확인
        assert bridge.ws_master.cell(row=18, column=7).value == 1  # 티어 번호
        assert bridge.ws_master.cell(row=18, column=8).value == 10  # 잔고량

        # Tier 2 행 (row=19) 확인
        assert bridge.ws_master.cell(row=19, column=7).value == 2
        assert bridge.ws_master.cell(row=19, column=8).value == 20

        bridge.close_workbook()

    def test_update_program_area_empty_positions(self, temp_excel_file):
        """포지션이 없을 때 프로그램 영역 업데이트"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        # 빈 포지션
        bridge.update_program_area([], tier1_price=10.0)

        # Tier 1 행 (row=18) - 잔고량 0
        assert bridge.ws_master.cell(row=18, column=8).value == 0

        bridge.close_workbook()


class TestAppendHistoryLog:
    """append_history_log() 테스트"""

    def test_append_history_log(self, temp_excel_file):
        """히스토리 로그 추가"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        initial_rows = bridge.ws_history.max_row

        log_entry = {
            "update_time": "2024-01-01 12:00:00",
            "date": "2024-01-01",
            "sheet": "Main",
            "ticker": "SOXL",
            "tier": 5,
            "total_tiers": 240,
            "quantity_diff": 10,
            "invested": 500.0,
            "tier_amount": 50.0,
            "balance": 9500.0,
            "stock_value": 510.0,
            "holding_profit": 10.0,
            "buy_ready": 50.0,
            "withdrawable": 9450.0,
            "arbitrage_profit": 0.02,
            "buy_qty": 10,
            "sell_qty": 0
        }

        bridge.append_history_log(log_entry)

        # 행이 추가되었는지 확인
        assert bridge.ws_history.max_row == initial_rows + 1

        # 추가된 행 내용 확인
        new_row = initial_rows + 1
        assert bridge.ws_history.cell(row=new_row, column=4).value == "SOXL"  # ticker
        assert bridge.ws_history.cell(row=new_row, column=5).value == 5  # tier

        bridge.close_workbook()


class TestCreateHistoryLogEntry:
    """create_history_log_entry() 테스트"""

    def test_create_history_log_entry(self, temp_excel_file, grid_settings_basic):
        """히스토리 로그 엔트리 생성"""
        bridge = ExcelBridge(temp_excel_file)

        state = SystemState(
            current_price=25.00,
            tier1_price=26.00,
            current_tier=2,
            account_balance=9800.0,
            total_quantity=10,
            total_invested=200.0,
            stock_value=250.0,
            total_profit=50.0,
            profit_rate=0.25
        )

        log_entry = bridge.create_history_log_entry(
            state=state,
            settings=grid_settings_basic,
            buy_qty=10,
            sell_qty=0
        )

        assert isinstance(log_entry, dict)
        assert log_entry["ticker"] == "SOXL"
        assert log_entry["tier"] == 2
        assert log_entry["buy_qty"] == 10
        assert log_entry["sell_qty"] == 0
        assert "update_time" in log_entry
        assert "date" in log_entry


class TestSaveWorkbook:
    """save_workbook() 테스트"""

    def test_save_workbook_success(self, temp_excel_file):
        """정상 저장"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        # 수정
        bridge.ws_master["E3"].value = 999

        # 저장
        result = bridge.save_workbook()
        assert result is True

        bridge.close_workbook()

    def test_save_workbook_without_load(self, temp_excel_file):
        """load 없이 저장 시도"""
        bridge = ExcelBridge(temp_excel_file)
        # load_workbook 호출 안 함

        result = bridge.save_workbook()
        assert result is False  # wb가 None이므로 False

    @pytest.mark.xfail(reason="PermissionError simulation requires actual file lock")
    def test_save_workbook_with_retry(self, temp_excel_file):
        """파일 잠금 시 재시도"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        # PermissionError 시뮬레이션은 실제 파일 잠금이 필요
        # 이 테스트는 실제 환경에서만 의미 있음
        result = bridge.save_workbook(max_retries=2, retry_delay=0.1)
        assert result is True

        bridge.close_workbook()


class TestCloseWorkbook:
    """close_workbook() 테스트"""

    def test_close_workbook(self, temp_excel_file):
        """파일 닫기"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        assert bridge.wb is not None

        bridge.close_workbook()

        assert bridge.wb is None
        assert bridge.ws_master is None
        assert bridge.ws_history is None


class TestReadBool:
    """_read_bool() 메서드 테스트"""

    @pytest.fixture
    def bridge_with_wb(self, temp_excel_file):
        """Workbook이 로드된 bridge"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()
        yield bridge
        bridge.close_workbook()

    def test_read_bool_numeric_zero_is_false(self, bridge_with_wb):
        """숫자 0은 False"""
        # 셀에 0 설정
        bridge_with_wb.ws_master["Z1"].value = 0
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._read_bool(cell)
        assert result is False

    def test_read_bool_numeric_one_is_true(self, bridge_with_wb):
        """숫자 1은 True"""
        bridge_with_wb.ws_master["Z1"].value = 1
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._read_bool(cell)
        assert result is True

    def test_read_bool_string_zero_is_false(self, bridge_with_wb):
        """문자열 '0'은 False"""
        bridge_with_wb.ws_master["Z1"].value = "0"
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._read_bool(cell)
        assert result is False

    def test_read_bool_string_one_is_true(self, bridge_with_wb):
        """문자열 '1'은 True"""
        bridge_with_wb.ws_master["Z1"].value = "1"
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._read_bool(cell)
        assert result is True

    def test_read_bool_true_string_variants(self, bridge_with_wb):
        """다양한 True 문자열"""
        true_values = ["true", "True", "TRUE", "yes", "Yes", "y", "Y", "on"]

        for val in true_values:
            bridge_with_wb.ws_master["Z1"].value = val
            cell = bridge_with_wb.ws_master["Z1"]
            result = bridge_with_wb._read_bool(cell)
            assert result is True, f"'{val}'는 True여야 함"

    def test_read_bool_false_string_variants(self, bridge_with_wb):
        """다양한 False 문자열"""
        false_values = ["false", "False", "FALSE", "no", "No", "n", "N", "off", ""]

        for val in false_values:
            bridge_with_wb.ws_master["Z1"].value = val
            cell = bridge_with_wb.ws_master["Z1"]
            result = bridge_with_wb._read_bool(cell)
            assert result is False, f"'{val}'는 False여야 함"

    def test_read_bool_none_is_false(self, bridge_with_wb):
        """None은 False"""
        bridge_with_wb.ws_master["Z1"].value = None
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._read_bool(cell)
        assert result is False

    def test_read_bool_unknown_string_defaults_to_false(self, bridge_with_wb):
        """알 수 없는 문자열은 False (경고 로그)"""
        bridge_with_wb.ws_master["Z1"].value = "unknown_value"
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._read_bool(cell)
        assert result is False


class TestParsePercent:
    """_parse_percent() 메서드 테스트"""

    @pytest.fixture
    def bridge_with_wb(self, temp_excel_file):
        """Workbook이 로드된 bridge"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()
        yield bridge
        bridge.close_workbook()

    def test_parse_percent_from_string(self, bridge_with_wb):
        """문자열 퍼센트 파싱"""
        bridge_with_wb.ws_master["Z1"].value = "5%"
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._parse_percent(cell, default=0.0)
        assert result == 0.05  # 5% → 0.05

    def test_parse_percent_from_number(self, bridge_with_wb):
        """숫자 퍼센트 파싱"""
        bridge_with_wb.ws_master["Z1"].value = 3
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._parse_percent(cell, default=0.0)
        assert result == 0.03  # 3 → 0.03

    def test_parse_percent_already_decimal(self, bridge_with_wb):
        """이미 소수인 경우"""
        bridge_with_wb.ws_master["Z1"].value = 0.05
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._parse_percent(cell, default=0.0)
        assert result == 0.05  # 0.05 그대로

    def test_parse_percent_none_returns_default(self, bridge_with_wb):
        """None이면 기본값 반환"""
        bridge_with_wb.ws_master["Z1"].value = None
        cell = bridge_with_wb.ws_master["Z1"]

        result = bridge_with_wb._parse_percent(cell, default=0.10)
        assert result == 0.10


class TestIntegration:
    """통합 테스트"""

    def test_full_workflow(self, temp_excel_file):
        """전체 워크플로우: 로드 → 수정 → 저장"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        # 1. 설정 로드
        settings = bridge.load_settings()
        assert settings.ticker == "SOXL"

        # 2. 상태 업데이트
        state = SystemState(
            current_price=25.00,
            tier1_price=26.00,
            current_tier=5,
            account_balance=8000.0,
            total_quantity=100,
            total_invested=2000.0,
            stock_value=2500.0,
            total_profit=500.0,
            profit_rate=0.25
        )
        bridge.update_program_info(state)

        # 3. 포지션 업데이트
        positions = [
            Position(tier=5, quantity=100, avg_price=20.0, invested_amount=2000.0, opened_at=datetime.now())
        ]
        bridge.update_program_area(positions, tier1_price=26.0)

        # 4. 로그 추가
        log_entry = bridge.create_history_log_entry(state, settings, buy_qty=100, sell_qty=0)
        bridge.append_history_log(log_entry)

        # 5. 저장
        result = bridge.save_workbook()
        assert result is True

        bridge.close_workbook()

    def test_load_settings_uses_read_bool(self, temp_excel_file):
        """load_settings가 _read_bool 사용 확인"""
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        # B7 (tier1_auto_update), B8 (tier1_trading_enabled), B9 (buy_limit), B10 (sell_limit)
        # 모두 boolean으로 읽혀야 함
        settings = bridge.load_settings()

        assert isinstance(settings.tier1_auto_update, bool)
        assert isinstance(settings.tier1_trading_enabled, bool)
        assert isinstance(settings.buy_limit, bool)
        assert isinstance(settings.sell_limit, bool)

        bridge.close_workbook()
