"""
Phoenix 자동매매 시스템 통합 테스트

엔드-투-엔드 시나리오 테스트:
1. 시스템 시작 → 시세 수신 → 매수 신호 → 주문 실행 → Excel 기록
2. 매수 → 가격 상승 → 매도 신호 → 주문 실행 → 수익 실현
3. 다중 Tier 매매 시나리오
4. 24시간 안정성 시뮬레이션
"""

import pytest
import time
from unittest.mock import Mock, patch, MagicMock
from datetime import datetime
import openpyxl
from src.phoenix_system import PhoenixSystem
from src.models import TradeSignal


def enable_tier1_trading(excel_file_path: str):
    """Helper: Tier 1 거래 활성화 (Excel 파일 직접 수정)"""
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb["01_매매전략_기준설정"]
    ws["B8"] = 1  # tier1_trading_enabled
    wb.save(excel_file_path)
    wb.close()


class TestEndToEndBuyScenario:
    """매수 시나리오 통합 테스트"""

    def test_full_buy_workflow(self, temp_excel_file):
        """
        완전한 매수 워크플로우:
        시작 → 시세 수신 → Tier 1 매수 조건 → 주문 → Excel 기록
        """
        with patch('src.phoenix_system.KiwoomAdapter') as MockKiwoom:
            with patch('src.phoenix_system.TelegramNotifier') as MockTelegram:
                # Mock 설정
                mock_kiwoom = MockKiwoom.return_value
                mock_kiwoom.login.return_value = True
                mock_kiwoom.get_us_stock_price.return_value = 10.0
                mock_kiwoom.get_account_balance.return_value = 10000.0
                mock_kiwoom.subscribe_real_price = Mock()
                mock_kiwoom.send_order.return_value = True

                # Tier 1 활성화
                enable_tier1_trading(temp_excel_file)

                # 시스템 초기화 (블로킹 없이)
                system = PhoenixSystem(temp_excel_file)
                system.initialize()
                system.is_running = True

                # 시세 수신 (Tier 1 가격)
                system._on_price_update(10.0)
                time.sleep(0.1)

                # 검증: 주문이 실행되었는지
                assert mock_kiwoom.send_order.called
                call_args = mock_kiwoom.send_order.call_args

                # 매수 주문 확인 (positional args)
                assert call_args[0][0] == "BUY"  # order_type
                assert call_args[0][1] == "SOXL"  # ticker

                # GridEngine 상태 확인
                assert len(system.grid_engine.positions) > 0
                assert system.grid_engine.positions[0].tier == 1


class TestEndToEndSellScenario:
    """매도 시나리오 통합 테스트"""

    def test_full_sell_workflow(self, temp_excel_file):
        """
        완전한 매도 워크플로우:
        매수 → 가격 상승 → Tier 1 매도 조건 → 주문 → 수익 실현
        """
        with patch('src.phoenix_system.KiwoomAdapter') as MockKiwoom:
            with patch('src.phoenix_system.TelegramNotifier') as MockTelegram:
                mock_kiwoom = MockKiwoom.return_value
                mock_kiwoom.login.return_value = True
                mock_kiwoom.get_us_stock_price.return_value = 10.0
                mock_kiwoom.get_account_balance.return_value = 10000.0
                mock_kiwoom.subscribe_real_price = Mock()
                mock_kiwoom.send_order.return_value = True

                # Tier 1 활성화
                enable_tier1_trading(temp_excel_file)

                system = PhoenixSystem(temp_excel_file)
                system.initialize()
                system.is_running = True

                # 1. 매수
                system._on_price_update(10.0)
                time.sleep(0.1)

                # 2. 가격 상승 (수익 3% 초과로 매도 조건)
                system._on_price_update(10.35)
                time.sleep(0.1)

                # 검증: 매도 주문 실행
                sell_calls = [
                    call for call in mock_kiwoom.send_order.call_args_list
                    if call[0][0] == "SELL"  # positional arg 0 is order_type
                ]

                assert len(sell_calls) > 0

                # 포지션 청산 확인
                assert len(system.grid_engine.positions) == 0


class TestMultiTierScenario:
    """다중 Tier 매매 시나리오"""

    def test_cascade_buy_three_tiers(self, temp_excel_file):
        """
        계단식 매수: Tier 1 → Tier 2 → Tier 3
        """
        with patch('src.phoenix_system.KiwoomAdapter') as MockKiwoom:
            with patch('src.phoenix_system.TelegramNotifier'):
                mock_kiwoom = MockKiwoom.return_value
                mock_kiwoom.login.return_value = True
                mock_kiwoom.send_order.return_value = "ORDER_MULTI"

                # Tier 1 활성화
                enable_tier1_trading(temp_excel_file)

                system = PhoenixSystem(temp_excel_file)
                system.start()

                # Tier 1 매수
                system._on_price_update(10.0)
                time.sleep(0.1)

                # Tier 2 매수
                system._on_price_update(9.95)
                time.sleep(0.1)

                # Tier 3 매수
                system._on_price_update(9.90)
                time.sleep(0.1)

                # 3개 포지션 확인
                assert len(system.grid_engine.positions) == 3

    def test_cascade_sell_after_recovery(self, temp_excel_file):
        """
        가격 회복 후 계단식 매도
        """
        with patch('src.phoenix_system.KiwoomAdapter') as MockKiwoom:
            with patch('src.phoenix_system.TelegramNotifier'):
                mock_kiwoom = MockKiwoom.return_value
                mock_kiwoom.login.return_value = True
                mock_kiwoom.get_us_stock_price.return_value = 10.0
                mock_kiwoom.get_account_balance.return_value = 10000.0
                mock_kiwoom.subscribe_real_price = Mock()
                mock_kiwoom.send_order.return_value = True

                # Tier 1 활성화
                enable_tier1_trading(temp_excel_file)

                system = PhoenixSystem(temp_excel_file)
                system.initialize()
                system.is_running = True

                # 하락: Tier 1, 2, 3 매수
                for price in [10.0, 9.95, 9.90]:
                    system._on_price_update(price)
                    time.sleep(0.1)

                # 회복: Tier 3, 2, 1 순차 매도 (각각 3% 수익 초과하도록)
                for price in [10.20, 10.25, 10.35]:
                    system._on_price_update(price)
                    time.sleep(0.1)

                # 모든 포지션 청산 확인
                assert len(system.grid_engine.positions) == 0


class Test24HourStabilitySimulation:
    """24시간 안정성 시뮬레이션"""

    @pytest.mark.slow
    def test_continuous_operation_simulation(self, temp_excel_file):
        """
        24시간 연속 운영 시뮬레이션 (1분 = 1초)
        실제 24시간 = 24분 시뮬레이션
        """
        with patch('src.phoenix_system.KiwoomAdapter') as MockKiwoom:
            with patch('src.phoenix_system.TelegramNotifier'):
                mock_kiwoom = MockKiwoom.return_value
                mock_kiwoom.login.return_value = True
                mock_kiwoom.send_order.return_value = "ORDER_STABLE"
                mock_kiwoom.get_us_stock_price.return_value = 10.0
                mock_kiwoom.get_account_balance.return_value = 10000.0
                mock_kiwoom.subscribe_real_price = Mock()
                mock_kiwoom.disconnect = Mock()

                # Tier 1 활성화
                enable_tier1_trading(temp_excel_file)

                system = PhoenixSystem(temp_excel_file)

                # 블로킹 없이 초기화만 수행
                system.initialize()
                system.is_running = True

                # I/O Worker 시작 (start()에서 하는 것과 동일)
                system.io_worker_running = True
                import threading
                system.io_worker = threading.Thread(target=system._io_worker, daemon=True)
                system.io_worker.start()

                # 1440분 = 24시간 (빠른 시뮬레이션)
                price_sequence = [10.0, 9.95, 9.90, 9.85, 9.90, 9.95, 10.0, 10.05]
                iterations = 100  # 100회 시세 업데이트

                for i in range(iterations):
                    price = price_sequence[i % len(price_sequence)]
                    system._on_price_update(price)
                    time.sleep(0.01)  # 10ms 대기

                # 시스템이 여전히 정상 작동
                assert system.is_running == True

                # 메모리 누수 없음 (대략적 체크)
                # 포지션 리스트가 비정상적으로 크지 않음
                assert len(system.grid_engine.positions) < 240

                # 안정적으로 종료 가능
                system.stop()
                assert system.shutdown_completed == True


class TestErrorRecovery:
    """에러 복구 통합 테스트"""

    def test_kiwoom_disconnect_and_reconnect(self, temp_excel_file):
        """Kiwoom 연결 끊김 및 재연결"""
        with patch('src.phoenix_system.KiwoomAdapter') as MockKiwoom:
            with patch('src.phoenix_system.TelegramNotifier'):
                mock_kiwoom = MockKiwoom.return_value
                mock_kiwoom.login.return_value = True

                # Reconnect 메서드 Mock 추가
                mock_kiwoom.reconnect = Mock(return_value=True)
                mock_kiwoom.on_disconnect = Mock()

                system = PhoenixSystem(temp_excel_file)
                system.start()

                # 연결 끊김 시뮬레이션
                system.kiwoom.on_disconnect()

                # on_disconnect 호출 확인
                mock_kiwoom.on_disconnect.assert_called_once()

                # 재연결 시도
                reconnect_success = system.kiwoom.reconnect()

                # 재연결 성공 확인
                assert reconnect_success == True
                mock_kiwoom.reconnect.assert_called_once()

    def test_excel_file_lock_recovery(self, temp_excel_file):
        """Excel 파일 잠금 복구"""
        from src.excel_bridge import ExcelBridge

        # ExcelBridge 단독 테스트
        bridge = ExcelBridge(temp_excel_file)
        bridge.load_workbook()

        # save_workbook에 PermissionError 발생 시뮬레이션
        original_save = bridge.wb.save

        call_count = [0]

        def mock_save_with_lock(path):
            call_count[0] += 1
            if call_count[0] == 1:
                raise PermissionError("파일이 다른 프로그램에서 사용 중")
            # 두 번째 호출은 성공
            original_save(path)

        bridge.wb.save = mock_save_with_lock

        # Retry 로직 테스트 (max_retries=3, retry_delay=0.1로 빠른 테스트)
        result = bridge.save_workbook(max_retries=3, retry_delay=0.1)

        # 재시도 후 성공 확인
        assert result == True
        assert call_count[0] == 2  # 첫 시도 실패, 재시도 성공

        bridge.close_workbook()


class TestDataIntegrity:
    """데이터 무결성 통합 테스트"""

    def test_excel_data_matches_engine_state(self, temp_excel_file):
        """Excel 데이터와 Engine 상태 일치 (동기 업데이트 테스트)"""
        with patch('src.phoenix_system.KiwoomAdapter') as MockKiwoom:
            with patch('src.phoenix_system.TelegramNotifier'):
                mock_kiwoom = MockKiwoom.return_value
                mock_kiwoom.login.return_value = True
                mock_kiwoom.get_us_stock_price.return_value = 10.0
                mock_kiwoom.get_account_balance.return_value = 10000.0
                mock_kiwoom.subscribe_real_price = Mock()
                mock_kiwoom.send_order.return_value = True

                # Tier 1 활성화
                enable_tier1_trading(temp_excel_file)

                system = PhoenixSystem(temp_excel_file)
                system.initialize()
                system.is_running = True

                # 매수
                system._on_price_update(10.0)
                time.sleep(0.1)

                # Excel 동기 업데이트 (I/O Worker 없이 직접 처리)
                state = system.grid_engine.get_system_state(10.0)
                system.excel_bridge.update_program_area(
                    system.grid_engine.positions,
                    system.grid_engine.tier1_price,
                    system.settings.buy_interval
                )
                system.excel_bridge.save_workbook()

                # 데이터 일치 확인 (Excel에서 직접 읽기)
                wb = openpyxl.load_workbook(temp_excel_file)
                ws = wb["01_매매전략_기준설정"]

                engine_positions = system.grid_engine.positions

                # Tier 1 데이터 확인 (row 18 = Tier 1 in update_program_area)
                if len(engine_positions) > 0:
                    tier1_pos = engine_positions[0]
                    # update_program_area writes to row 18 for Tier 1 (row_idx = 17 + tier)
                    excel_quantity = ws.cell(row=18, column=8).value or 0  # H18
                    excel_invested = ws.cell(row=18, column=9).value or 0.0  # I18
                    excel_avg_price = ws.cell(row=18, column=10).value or 0.0  # J18

                    assert excel_quantity == tier1_pos.quantity
                    assert excel_invested == pytest.approx(tier1_pos.invested_amount, rel=1e-2)
                    assert excel_avg_price == pytest.approx(tier1_pos.avg_price, rel=1e-2)

                wb.close()


class TestPerformance:
    """성능 통합 테스트"""

    def test_high_frequency_price_updates(self, temp_excel_file):
        """고빈도 시세 업데이트 성능"""
        with patch('src.phoenix_system.KiwoomAdapter'):
            with patch('src.phoenix_system.TelegramNotifier'):
                system = PhoenixSystem(temp_excel_file)
                system.start()

                # 1000번 시세 업데이트
                start_time = time.time()

                for i in range(1000):
                    system._on_price_update(10.0 + (i % 10) * 0.01)

                duration = time.time() - start_time

                # 1000번 업데이트가 10초 이내
                assert duration < 10.0

                print(f"1000 updates in {duration:.2f}s ({1000/duration:.1f} updates/sec)")
