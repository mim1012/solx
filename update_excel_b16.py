# -*- coding: utf-8 -*-
"""Update Excel template to add price_check_interval setting"""
import openpyxl

# Load workbook
wb = openpyxl.load_workbook('phoenix_grid_template_v3.xlsx')
ws = wb['01_매매전략_기준설정']

# Add price check interval setting to B16
ws['A16'] = '시세 조회 주기 (초)'
ws['B16'] = 40.0

# Save workbook
wb.save('phoenix_grid_template_v3.xlsx')
wb.close()

print("✅ Excel 템플릿 업데이트 완료:")
print("   A16: '시세 조회 주기 (초)'")
print("   B16: 40.0")
